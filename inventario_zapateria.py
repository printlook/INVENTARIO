import sqlite3
import tkinter as tk
from tkinter import messagebox, messagebox, ttk, filedialog
import os
from datetime import datetime
import hashlib
from tkinter import filedialog
import shutil
from openpyxl import Workbook
import openpyxl 
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import xlsxwriter
import locale

from tkcalendar import Calendar,DateEntry
            
PRIMARY_COLOR = "#2c3e50"
SECONDARY_COLOR = "#ecf0f1"
ACCENT_COLOR = "#e74c3c"
FONT_COLOR = "#ecf0f1"
FONT = ("Arial", 12)
import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
import hashlib
import os

class LoginSystem:
    def __init__(self):
        # Crear directorio de base de datos si no existe
        if not os.path.exists('databases'):
            os.makedirs('databases')
        
        # Conectar o crear base de datos de usuarios
        self.conn = sqlite3.connect('databases/users.db')
        self.cursor = self.conn.cursor()
        
        # Crear tabla de usuarios si no existe
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        self.conn.commit()

    def hash_password(self, password):
        """Encripta la contraseña usando SHA-256"""
        return hashlib.sha256(password.encode()).hexdigest()

    def validate_admin_password(self, password):
        """Valida la contraseña de administrador"""
        return password == "384916139"

    def create_user(self, username, password):
        """Crea un nuevo usuario en la base de datos"""
        try:
            hashed_password = self.hash_password(password)
            self.cursor.execute(
                "INSERT INTO usuarios (username, password) VALUES (?, ?)",
                (username, hashed_password)
            )
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False

    def verify_user(self, username, password):
        """Verifica las credenciales del usuario"""
        hashed_password = self.hash_password(password)
        self.cursor.execute(
            "SELECT * FROM usuarios WHERE username = ? AND password = ?",
            (username, hashed_password)
        )
        return self.cursor.fetchone() is not None

class LoginApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Login")
        
        # Centrar la ventana en la pantalla
        window_width = 400
        window_height = 500
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        center_x = int((screen_width - window_width) / 2)
        center_y = int((screen_height - window_height) / 2)
        self.root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
        self.root.resizable(False, False)
        
        # Variable para controlar la visibilidad de la contraseña
        self.show_password = tk.BooleanVar()
        
        # Configurar estilo
        self.style = ttk.Style()
        self.style.configure("Custom.TFrame", 
                           background="#f8f9fa")
        
        self.style.configure("Title.TLabel", 
                           font=("Helvetica", 24, "bold"), 
                           background="#f8f9fa",
                           foreground="#2c3e50")
        
        self.style.configure("Field.TLabel", 
                           font=("Helvetica", 12),
                           background="#f8f9fa",
                           foreground="#495057")
        
        self.style.configure("Custom.TButton", 
                           font=("Helvetica", 12),
                           foreground="black")  # Texto negro en botones
        
        self.login_system = LoginSystem()
        self.setup_ui()
        self.root.bind('<Return>', lambda e: self.login())

    def toggle_password(self):
        if self.show_password.get():
            self.password_entry.config(show="")
            self.toggle_btn.config(text="🔒")
        else:
            self.password_entry.config(show="•")
            self.toggle_btn.config(text="👁️")

    def setup_ui(self):
        main_frame = ttk.Frame(self.root, style="Custom.TFrame", padding="30")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Título
        title_label = ttk.Label(main_frame, 
                              text="Bienvenido", 
                              style="Title.TLabel")
        title_label.pack(pady=(0, 30))

        # Frame de login
        self.login_frame = ttk.Frame(main_frame, style="Custom.TFrame")
        self.login_frame.pack(fill=tk.BOTH, expand=True)

        # Campo de usuario con ícono
        user_frame = ttk.Frame(self.login_frame, style="Custom.TFrame")
        user_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(user_frame, 
                 text="👤", 
                 font=("Segoe UI Emoji", 12),
                 style="Field.TLabel").pack(side=tk.LEFT, padx=(0, 10))
        
        self.username_entry = ttk.Entry(user_frame, 
                                      width=30,
                                      font=("Helvetica", 11))
        self.username_entry.pack(side=tk.LEFT, expand=True)
        self.username_entry.bind('<Return>', lambda e: self.password_entry.focus())

        # Campo de contraseña con ícono y botón de visibilidad
        pass_frame = ttk.Frame(self.login_frame, style="Custom.TFrame")
        pass_frame.pack(fill=tk.X, pady=(0, 30))
        
        ttk.Label(pass_frame, 
                 text="🔒", 
                 font=("Segoe UI Emoji", 12),
                 style="Field.TLabel").pack(side=tk.LEFT, padx=(0, 10))
        
        self.password_entry = ttk.Entry(pass_frame, 
                                      show="•", 
                                      width=30,
                                      font=("Helvetica", 11))
        self.password_entry.pack(side=tk.LEFT, expand=True)
        
        self.toggle_btn = tk.Button(pass_frame, 
                                  text="👁️",
                                  font=("Segoe UI Emoji", 10),
                                  command=lambda: self.show_password.set(not self.show_password.get()),
                                  bd=0,
                                  bg="#f8f9fa",
                                  cursor="hand2")
        self.toggle_btn.pack(side=tk.RIGHT, padx=(5, 0))
        
        # Vincular el cambio de visibilidad de contraseña
        self.show_password.trace('w', lambda *args: self.toggle_password())

        # Botones con texto negro
        self.login_button = ttk.Button(self.login_frame, 
                                     text="Iniciar Sesión",
                                     command=self.login,
                                     style="Custom.TButton",
                                     width=25)
        self.login_button.pack(pady=(0, 15))
        
        self.register_button = ttk.Button(self.login_frame,
                                        text="Registrar Nuevo Usuario",
                                        command=self.show_register_window,
                                        style="Custom.TButton",
                                        width=25)
        self.register_button.pack()

    def show_register_window(self):
        register_window = tk.Toplevel(self.root)
        register_window.title("Registrar Nuevo Usuario")
        
        # Centrar ventana de registro
        window_width = 400
        window_height = 500
        screen_width = register_window.winfo_screenwidth()
        screen_height = register_window.winfo_screenheight()
        center_x = int((screen_width - window_width) / 2)
        center_y = int((screen_height - window_height) / 2)
        register_window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
        
        register_window.transient(self.root)
        register_window.grab_set()

        # Frame principal
        frame = ttk.Frame(register_window, style="Custom.TFrame", padding="30")
        frame.pack(fill=tk.BOTH, expand=True)

        # Título
        ttk.Label(frame, 
                 text="Registro de Usuario", 
                 style="Title.TLabel").pack(pady=(0, 30))

        # Campos de registro con íconos
        admin_frame = ttk.Frame(frame, style="Custom.TFrame")
        admin_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(admin_frame, 
                 text="🔑",
                 font=("Segoe UI Emoji", 12)).pack(side=tk.LEFT, padx=(0, 10))
        
        admin_password_entry = ttk.Entry(admin_frame, 
                                       show="•", 
                                       width=30,
                                       font=("Helvetica", 11))
        admin_password_entry.pack(fill=tk.X)

        user_frame = ttk.Frame(frame, style="Custom.TFrame")
        user_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(user_frame, 
                 text="👤",
                 font=("Segoe UI Emoji", 12)).pack(side=tk.LEFT, padx=(0, 10))
        
        new_username_entry = ttk.Entry(user_frame, 
                                     width=30,
                                     font=("Helvetica", 11))
        new_username_entry.pack(fill=tk.X)

        pass_frame = ttk.Frame(frame, style="Custom.TFrame")
        pass_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(pass_frame, 
                 text="🔒",
                 font=("Segoe UI Emoji", 12)).pack(side=tk.LEFT, padx=(0, 10))
        
        new_password_entry = ttk.Entry(pass_frame, 
                                     show="•", 
                                     width=30,
                                     font=("Helvetica", 11))
        new_password_entry.pack(fill=tk.X)

        confirm_frame = ttk.Frame(frame, style="Custom.TFrame")
        confirm_frame.pack(fill=tk.X, pady=(0, 30))
        
        ttk.Label(confirm_frame, 
                 text="🔒",
                 font=("Segoe UI Emoji", 12)).pack(side=tk.LEFT, padx=(0, 10))
        
        confirm_password_entry = ttk.Entry(confirm_frame, 
                                         show="•", 
                                         width=30,
                                         font=("Helvetica", 11))
        confirm_password_entry.pack(fill=tk.X)

        def register():
            admin_password = admin_password_entry.get()
            username = new_username_entry.get().strip()
            password = new_password_entry.get()
            confirm_password = confirm_password_entry.get()

            if not all([admin_password, username, password, confirm_password]):
                messagebox.showerror("Error", "Todos los campos son obligatorios")
                return

            if not self.login_system.validate_admin_password(admin_password):
                messagebox.showerror("Error", "Contraseña de administrador incorrecta")
                return

            if password != confirm_password:
                messagebox.showerror("Error", "Las contraseñas no coinciden")
                return

            if len(password) < 6:
                messagebox.showerror("Error", "La contraseña debe tener al menos 6 caracteres")
                return

            if self.login_system.create_user(username, password):
                messagebox.showinfo("Éxito", "Usuario registrado correctamente")
                register_window.destroy()
            else:
                messagebox.showerror("Error", "El nombre de usuario ya existe")

        # Botón de registro con texto negro
        register_button = ttk.Button(frame, 
                                   text="Registrar", 
                                   command=register,
                                   style="Custom.TButton",
                                   width=25)
        register_button.pack(pady=(0, 10))

    def login(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get()

        if not username or not password:
            messagebox.showerror("Error", "Por favor ingrese usuario y contraseña")
            return

        if self.login_system.verify_user(username, password):
            messagebox.showinfo("Éxito", "Inicio de sesión exitoso")
            self.root.destroy()
            root = tk.Tk()
            app = InventoryApp(root)  # Asumiendo que tienes una clase InventoryApp
            root.mainloop()
        else:
            messagebox.showerror("Error", "Usuario o contraseña incorrectos")


    
class InventoryManager:
    def __init__(self):
        self.current_db = None
        self.connection = None
        self.cursor = None
    
    def create_connection(self, db_name):
        """Crea o conecta a una base de datos específica"""
        try:
            # Cerrar conexión existente si hay una
            if self.connection:
                self.connection.close()
            
            # Crear la carpeta databases si no existe
            if not os.path.exists('databases'):
                os.makedirs('databases')
            
            # Conectar a la nueva base de datos
            self.current_db = db_name
            self.connection = sqlite3.connect(f"databases/{db_name}.db")
            self.cursor = self.connection.cursor()
            
            # Crear tablas si no existen
            self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS productos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                codigo TEXT NOT NULL UNIQUE,
                precio_costo REAL NOT NULL,
                precio_venta REAL NOT NULL,
                cantidad INTEGER NOT NULL,
                cantidad_inicial INTEGER NOT NULL,
                vitrina INTEGER DEFAULT 0,  -- Nueva columna
                bodega INTEGER DEFAULT 0,   -- Nueva columna
                descripcion TEXT,
                categoria TEXT,
                fecha_actualizacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
            
            self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS categorias (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL UNIQUE
            )
            ''')
            
            self.connection.commit()
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Error al conectar a la base de datos: {str(e)}")
            return False

    def get_available_databases(self):
        """Obtiene lista de bases de datos disponibles"""
        if not os.path.exists('databases'):
            return []
        return [f[:-3] for f in os.listdir('databases') if f.endswith('.db')]

    def backup_database(self):
        """Realiza una copia de seguridad de la base de datos actual"""
        if not self.current_db:
            return False
        
        try:
            if not os.path.exists('backups'):
                os.makedirs('backups')
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = f"backups/{self.current_db}_backup_{timestamp}.db"
            
            self.connection.close()
            shutil.copy2(f"databases/{self.current_db}.db", backup_path)
            self.create_connection(self.current_db)
            
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Error al crear backup: {str(e)}")
            return False

    def restore_database(self, backup_file):
        """Restaura una base de datos desde un backup"""
        try:
            if self.connection:
                self.connection.close()
            
            db_name = backup_file.split('_backup_')[0]
            shutil.copy2(f"backups/{backup_file}", f"databases/{db_name}.db")
            
            return self.create_connection(db_name)
        except Exception as e:
            messagebox.showerror("Error", f"Error al restaurar backup: {str(e)}")
            return False


class InventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema Multi-Inventario")
        self.manager = InventoryManager()
        self.root.geometry("800x600")
        self.root.resizable(True, True)
        self.setup_styles()
        self.setup_ui()
        self.sale_items = []  # Lista de productos en la venta
        self.ADMIN_PASSWORD = "15243"
        self.tree = None
        self.dialog = None

    def setup_styles(self):
       
        style = ttk.Style()
        style.configure("Treeview", rowheight=25)
        style.configure("Treeview.Heading", font=('Arial', 10, 'bold'))
    def cerrar_trimestre(self):
     """Actualiza el inventario inicial con la suma de vitrina y bodega."""
     if not self.validate_admin_password():
        return
    
     if messagebox.askyesno("Confirmar", "¿Desea cerrar el trimestre y actualizar el inventario inicial?"):
        try:
            # Actualizar la cantidad inicial con la suma de vitrina y bodega
            self.manager.cursor.execute('''
                UPDATE productos
                SET cantidad_inicial = vitrina + bodega
            ''')
            
            # Actualizar la cantidad actual con la suma de vitrina y bodega
            self.manager.cursor.execute('''
                UPDATE productos
                SET cantidad = vitrina + bodega
            ''')
            
            self.manager.connection.commit()
            messagebox.showinfo("Éxito", "El inventario inicial y la cantidad actual han sido actualizados con éxito.")
            
            # Actualizar la vista de la tabla principal
            self.load_inventory_data()
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al actualizar el inventario inicial: {str(e)}")

    
    def cuadre(self):
     """Muestra una ventana con el cuadre de inventario y permite modificar vitrina y bodega."""
     if not self.validate_admin_password():
            return
     if not self.manager.current_db:
        messagebox.showerror("Error", "No hay una base de datos seleccionada")
        return

     if self.dialog is not None:
        try:
            self.dialog.destroy()
        except:
            pass

     self.dialog = tk.Toplevel(self.root)
     self.dialog.title("Cuadre de Inventario")
     self.dialog.geometry("1400x700")
     self.dialog.transient(self.root)

     main_frame = ttk.Frame(self.dialog, padding="10")
     main_frame.pack(fill=tk.BOTH, expand=True)

    # Columnas de la tabla
     columns = (
        'Código', 'Nombre', 'Inventario Inicial', 'Ventas', 'Compras', 'Devoluciones',
        'Inventario Final', 'Vitrina', 'Bodega', 'Diferencia'
    )
    
     self.tree = ttk.Treeview(main_frame, columns=columns, show='headings', height=20)

    # Configurar columnas
     column_widths = {
        'Código': 100,
        'Nombre': 200,
        'Inventario Inicial': 120,
        'Ventas': 80,
        'Compras': 80,
        'Devoluciones': 100,
        'Inventario Final': 120,
        'Vitrina': 80,
        'Bodega': 80,
        'Diferencia': 100
    }

     for col in columns:
        self.tree.heading(col, text=col)
        self.tree.column(col, width=column_widths.get(col, 100))

    # Agregar scrollbar
     scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=self.tree.yview)
     scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
     self.tree.configure(yscrollcommand=scrollbar.set)
     self.tree.pack(fill=tk.BOTH, expand=True, pady=10)

     button_frame = ttk.Frame(main_frame)
     button_frame.pack(fill=tk.X, pady=10)

     # Agregar botón de actualizar
     actualizar_btn = ttk.Button(button_frame, text="Actualizar Datos", 
                               command=lambda: self.cargar_datos_cuadre())
     actualizar_btn.pack(side=tk.LEFT, padx=5)

     exportar_btn = ttk.Button(button_frame, text="Exportar a Excel", 
                             command=lambda: self.exportar_cuadre_a_excel())
     exportar_btn.pack(side=tk.LEFT, padx=5)

     modificar_btn = ttk.Button(button_frame, text="Modificar Vitrina y Bodega", 
                              command=lambda: self.modificar_vitrina_bodega())
     modificar_btn.pack(side=tk.LEFT, padx=5)
    

     generar_btn = ttk.Button(button_frame, text="Generar Hoja de Inventario", 
                            command=lambda: self.generar_hoja_inventario())
     generar_btn.pack(side=tk.LEFT, padx=5)
     reporte_precios_btn = ttk.Button(button_frame, text="Reporte de Precios", 
                               command=lambda: self.generar_reporte_precios())
     reporte_precios_btn.pack(side=tk.LEFT, padx=5)

     def on_closing():
        try:
            self.dialog.destroy()
        except:
            pass
        self.dialog = None

     self.dialog.protocol("WM_DELETE_WINDOW", on_closing)
    
    # Mostrar mensaje inicial
     self.tree.insert('', tk.END, values=("", "Presione 'Actualizar Datos' para cargar la información", "", "", "", "", "", "", "", ""))
    def generar_hoja_inventario(self):
      """Genera una hoja de inventario en Excel con campos vacíos para conteo manual."""
      try:
        # Crear un nuevo libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja de Inventario"

        # Agregar título y campos de información
        ws['B1'] = "HOJA DE INVENTARIO"
        ws['B1'].font = Font(bold=True, size=14)
        ws.merge_cells('B1:D1')  # Fusionar celdas para el título
        
        # Agregar campos de información
        ws['B3'] = "Nombre:"
        ws['B4'] = "Fecha:"
        ws['B5'] = "Encargado:"
        
        # Agregar líneas para llenar
        ws['C3'] = "_" * 30
        ws['C4'] = "_" * 30
        ws['C5'] = "_" * 30
        
        # Fusionar celdas para las líneas
        ws.merge_cells('C3:D3')
        ws.merge_cells('C4:D4')
        ws.merge_cells('C5:D5')

        # Definir encabezados de la tabla
        headers = ['Código', 'Nombre', 'Vitrina', 'Bodega', 'Total']
        
        # Escribir encabezados de la tabla (comenzando en la fila 7)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num, value=header)
            cell.font = Font(bold=True)
            
        # Obtener productos de la base de datos
        self.manager.cursor.execute('''
            SELECT codigo, nombre
            FROM productos
            ORDER BY nombre
        ''')
        productos = self.manager.cursor.fetchall()

        # Escribir productos y dejar campos vacíos para conteo
        for row_num, producto in enumerate(productos, 8):  # Comenzar en la fila 8
            ws.cell(row=row_num, column=1, value=producto[0])  # Código
            ws.cell(row=row_num, column=2, value=producto[1])  # Nombre
            # Las columnas de Vitrina, Bodega y Total se dejan vacías

        # Ajustar anchos de columna
        ws.column_dimensions[get_column_letter(1)].width = 10  # Código
        ws.column_dimensions[get_column_letter(2)].width = 15  # Nombre
        ws.column_dimensions[get_column_letter(3)].width = 30  # Vitrina
        ws.column_dimensions[get_column_letter(4)].width = 30  # Bodega
        ws.column_dimensions[get_column_letter(5)].width = 10  # Total

        # Agregar bordes a todas las celdas de la tabla
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aplicar bordes solo a la tabla de datos
        for row in ws.iter_rows(min_row=7, max_row=len(productos) + 7, max_col=5):
            for cell in row:
                cell.border = thin_border

        # Centrar el título
        ws['B1'].alignment = openpyxl.styles.Alignment(horizontal='center')

        # Guardar el archivo
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Guardar hoja de inventario"
        )

        if filename:
            wb.save(filename)
            messagebox.showinfo("Éxito", f"Hoja de inventario generada correctamente como {filename}")

      except Exception as e:
        messagebox.showerror("Error", f"Error al generar hoja de inventario: {str(e)}")

    def cargar_datos_cuadre(self):
     """Carga los datos del cuadre de inventario en la tabla."""
     try:
        # Limpiar la tabla
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Obtener el inventario inicial, vitrina y bodega
        self.manager.cursor.execute('''
            SELECT codigo, nombre, cantidad_inicial, vitrina, bodega
            FROM productos
        ''')
        productos = self.manager.cursor.fetchall()

        # Obtener las ventas
        self.manager.cursor.execute('''
            SELECT codigo_producto, SUM(cantidad)
            FROM detalle_ventas
            GROUP BY codigo_producto
        ''')
        ventas = {row[0]: row[1] for row in self.manager.cursor.fetchall()}

        # Obtener las compras
        self.manager.cursor.execute('''
            SELECT codigo_producto, SUM(cantidad)
            FROM detalle_compras
            GROUP BY codigo_producto
        ''')
        compras = {row[0]: row[1] for row in self.manager.cursor.fetchall()}

        # Obtener las devoluciones
        self.manager.cursor.execute('''
            SELECT codigo_producto, SUM(cantidad)
            FROM detalle_devolucion
            GROUP BY codigo_producto
        ''')
        devoluciones = {row[0]: row[1] for row in self.manager.cursor.fetchall()}

        # Calcular y mostrar datos
        for producto in productos:
            codigo, nombre, cantidad_inicial, vitrina, bodega = producto
            ventas_producto = ventas.get(codigo, 0) or 0
            compras_producto = compras.get(codigo, 0) or 0
            devoluciones_producto = devoluciones.get(codigo, 0) or 0

            # Calcular inventario final
            inventario_final = cantidad_inicial - ventas_producto + compras_producto - devoluciones_producto

            # Calcular la diferencia
            diferencia = inventario_final - (vitrina + bodega)

            # Mostrar la diferencia con signo
            diferencia_str = f"-{diferencia}" if diferencia >= 0 else f"+{abs(diferencia)}"

            # Insertar en la tabla
            self.tree.insert('', tk.END, values=(
                codigo,
                nombre,
                cantidad_inicial,
                ventas_producto,
                compras_producto,
                devoluciones_producto,
                inventario_final,
                vitrina,
                bodega,
                diferencia_str
            ))

     except Exception as e:
        messagebox.showerror("Error", f"Error al cargar el cuadre de inventario: {str(e)}")

    def modificar_vitrina_bodega(self):
     """Permite modificar las cantidades de vitrina y bodega en la tabla."""
     if not self.tree:
        return

     seleccion = self.tree.selection()
     if not seleccion:
        messagebox.showwarning("Advertencia", "Seleccione un producto para modificar")
        return

     item = self.tree.item(seleccion[0])
     valores = item['values']
     codigo = valores[0]

    # Crear ventana de modificación
     mod_dialog = tk.Toplevel(self.dialog)
     mod_dialog.title("Modificar Vitrina y Bodega")
     mod_dialog.geometry("300x150")
     mod_dialog.transient(self.dialog)

     frame = ttk.Frame(mod_dialog, padding="10")
     frame.pack(fill=tk.BOTH, expand=True)

     ttk.Label(frame, text="Vitrina:").grid(row=0, column=0, pady=5)
     vitrina_entry = ttk.Entry(frame, width=20)
     vitrina_entry.grid(row=0, column=1, pady=5)
     vitrina_entry.insert(0, valores[7])

     ttk.Label(frame, text="Bodega:").grid(row=1, column=0, pady=5)
     bodega_entry = ttk.Entry(frame, width=20)
     bodega_entry.grid(row=1, column=1, pady=5)
     bodega_entry.insert(0, valores[8])

     def guardar_cambios():
        try:
            vitrina = int(vitrina_entry.get())
            bodega = int(bodega_entry.get())

            if vitrina < 0 or bodega < 0:
                messagebox.showerror("Error", "Las cantidades no pueden ser negativas")
                return

            self.manager.cursor.execute('''
                UPDATE productos
                SET vitrina = ?, bodega = ?
                WHERE codigo = ?
            ''', (vitrina, bodega, codigo))

            self.manager.connection.commit()
            mod_dialog.destroy()
            self.cargar_datos_cuadre()
            messagebox.showinfo("Éxito", "Cantidades actualizadas correctamente")

        except ValueError:
            messagebox.showerror("Error", "Las cantidades deben ser números enteros")
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar las cantidades: {str(e)}")

     ttk.Button(frame, text="Guardar", command=guardar_cambios).grid(row=2, column=0, columnspan=2, pady=10)
    
    def generar_reporte_precios(self):
     """Genera un reporte de precios con costos totales."""
     try:
        # Crear un nuevo libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Precios"

        # Configurar el título y campos de información
        ws['B1'] = "INVENTARIO AL "
        ws['C1'] = "_" * 20  # Espacio para escribir la fecha
        ws['B1'].font = Font(bold=True, size=14)
        ws['C1'].font = Font(size=14)
        
        # Agregar nombre de base de datos
        ws['B2'] = ""
        ws['C2'] = "_" * 20  # Espacio para escribir el nombre
        cell = ws['B2']
        cell.font = Font(bold=True)

        # Aplicar fondo amarillo a las primeras filas
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=4):
            for cell in row:
                cell.fill = yellow_fill

        # Definir encabezados
        headers = ['Código', 'Nombre del artículo', 'Costos', 'Total de pago', 'Costo Total']
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = Font(bold=True)
            
        # Obtener datos de productos
        self.manager.cursor.execute('''
            SELECT codigo, nombre, precio_costo, (vitrina + bodega) as cantidad_total
            FROM productos
            ORDER BY nombre
        ''')
        productos = self.manager.cursor.fetchall()

        # Variables para totales
        total_cantidad = 0
        total_costo = 0
        total_precio_costo = 0

        # Escribir productos
        for row_num, producto in enumerate(productos, 5):
            codigo, nombre, precio_costo, cantidad = producto
            costo_total = precio_costo * cantidad if precio_costo and cantidad else 0
            
            total_cantidad += cantidad
            total_costo += costo_total
            total_precio_costo += precio_costo
            
            ws.cell(row=row_num, column=1, value=codigo)
            ws.cell(row=row_num, column=2, value=nombre)
            ws.cell(row=row_num, column=3, value=precio_costo)  # Precio costo
            ws.cell(row=row_num, column=4, value=cantidad)      # Cantidad total
            ws.cell(row=row_num, column=5, value=costo_total)   # Costo total (precio_costo * cantidad)

        # Escribir totales al final
        ultima_fila = len(productos) + 5
        cell_total_label = ws.cell(row=ultima_fila, column=2, value="TOTAL:")
        cell_total_label.font = Font(bold=True)
        
        # Total de precio costo
        cell_total_precio_costo = ws.cell(row=ultima_fila, column=3, value=total_precio_costo)
        cell_total_precio_costo.font = Font(bold=True)
        
        # Total de cantidad
        cell_total_cantidad = ws.cell(row=ultima_fila, column=4, value=total_cantidad)
        cell_total_cantidad.font = Font(bold=True)
        
        # Total de costo
        cell_total_costo = ws.cell(row=ultima_fila, column=5, value=total_costo)
        cell_total_costo.font = Font(bold=True)
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 10  # Código
        ws.column_dimensions['B'].width = 25  # Nombre
        ws.column_dimensions['C'].width = 10  # Costos
        ws.column_dimensions['D'].width = 10  # Total de pago
        ws.column_dimensions['E'].width = 10  # Costo Total

        # Agregar bordes a la tabla
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=4, max_row=ultima_fila, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border

        # Guardar el archivo
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Guardar reporte de precios"
        )

        if filename:
            wb.save(filename)
            messagebox.showinfo("Éxito", f"Reporte de precios generado correctamente como {filename}")

     except Exception as e:
        messagebox.showerror("Error", f"Error al generar reporte de precios: {str(e)}")
    
    def exportar_cuadre_a_excel(self):
     """Exporta el cuadre de inventario a un archivo Excel."""
     try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cuadre de Inventario"

        # Escribir encabezados
        columns = (
            'Código', 'Nombre', 'Inventario Inicial', 'Ventas', 'Compras', 'Devoluciones',
            'Inventario Final', 'Vitrina', 'Bodega', 'Diferencia'
        )
        for col_num, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=col_name).font = Font(bold=True)

        # Escribir datos
        for row_num, item in enumerate(self.tree.get_children(), 2):
            values = self.tree.item(item)['values']
            for col_num, value in enumerate(values, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Ajustar anchos
        for col in ws.columns:
            max_length = 0
            column = [cell for cell in col]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Guardar archivo
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Guardar cuadre de inventario"
        )

        if filename:
            wb.save(filename)
            messagebox.showinfo("Éxito", f"Reporte exportado correctamente como {filename}")

     except Exception as e:
        messagebox.showerror("Error", f"Error al exportar a Excel: {str(e)}")

    
    
        
    def export_to_report(self):
       """Exporta el historial a un archivo Excel con formato mejorado"""
       try:
        
        
        # Obtener datos completos del historial con detalles
        self.manager.cursor.execute('''
            SELECT 
                h.id as "ID Devolución",
                h.numero_envio as "Número de Envío",
                h.fecha_envio as "Fecha de Envío",
                h.fecha_devolucion as "Fecha de Devolución",
                d.codigo_producto as "Código de Producto",
                p.nombre as "Nombre de Producto",
                d.cantidad as "Cantidad Devuelta",
                d.precio_unitario as "Precio Unitario",
                d.subtotal as "Subtotal",
                h.total_devoluciones as "Total de Devolución"
            FROM historial_devoluciones h
            JOIN detalle_devolucion d ON h.id = d.devolucion_id
            JOIN productos p ON d.codigo_producto = p.codigo
            ORDER BY h.fecha_devolucion DESC, h.id, d.codigo_producto
        ''')
        
        # Obtener los datos y convertirlos a DataFrame
        data = self.manager.cursor.fetchall()
        columns = [description[0] for description in self.manager.cursor.description]
        df = pd.DataFrame(data, columns=columns)
        
        # Formatear las columnas de fecha
        df["Fecha de Envío"] = pd.to_datetime(df["Fecha de Envío"]).dt.strftime('%d/%m/%Y')
        df["Fecha de Devolución"] = pd.to_datetime(df["Fecha de Devolución"]).dt.strftime('%d/%m/%Y %H:%M:%S')
        
        # Formatear las columnas de moneda
        df["Precio Unitario"] = df["Precio Unitario"].apply(lambda x: f"Q{float(x):.2f}")
        df["Subtotal"] = df["Subtotal"].apply(lambda x: f"Q{float(x):.2f}")
        df["Total de Devolución"] = df["Total de Devolución"].apply(lambda x: f"Q{float(x):.2f}")
        
        # Generar nombre de archivo con fecha y hora
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"historial_devoluciones_{timestamp}.xlsx"
        
        # Crear un writer de Excel con el motor xlsxwriter
        with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
            # Escribir el DataFrame al archivo Excel
            df.to_excel(writer, sheet_name='Historial', index=False)
            
            # Obtener el objeto workbook y la hoja
            workbook = writer.book
            worksheet = writer.sheets['Historial']
            
            # Definir formatos
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'bg_color': '#D9D9D9',
                'border': 1
            })
            
            # Aplicar formato a los encabezados
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
                
            # Ajustar el ancho de las columnas
            for i, column in enumerate(df.columns):
                column_width = max(
                    df[column].astype(str).apply(len).max(),
                    len(column)
                )
                worksheet.set_column(i, i, column_width + 2)
        
        messagebox.showinfo("Éxito", f"Reporte exportado correctamente como '{filename}'")
        
       except ImportError:
        messagebox.showerror("Error", "Se requiere la librería pandas y xlsxwriter para exportar a Excel. " 
                            "Por favor, instale las librerías con:\n"
                            "pip install pandas xlsxwriter")
       except Exception as e:
        messagebox.showerror("Error", f"Error al exportar el historial: {str(e)}")  
        self.export_to_report()  
    
    def view_sales_history(self):
        """Muestra el historial de ventas en una ventana nueva."""
        if not self.manager.current_db:
            messagebox.showerror("Error", "No hay una base de datos seleccionada")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Historial de Ventas")
        dialog.geometry("1000x600")
        dialog.transient(self.root)
        dialog.grab_set()

        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
    
        # Frame para búsqueda y filtros
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=5)

        # Frame para el filtro de fechas
        date_frame = ttk.Frame(main_frame)
        date_frame.pack(fill=tk.X, pady=5)

        ttk.Label(date_frame, text="Fecha inicial:").pack(side=tk.LEFT, padx=5)
        start_date = DateEntry(date_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
        start_date.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(date_frame, text="Fecha final:").pack(side=tk.LEFT, padx=5)
        end_date = DateEntry(date_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
        end_date.pack(side=tk.LEFT, padx=5)

        ttk.Label(search_frame, text="Buscar por número de hoja de venta:").pack(side=tk.LEFT, padx=5)
        search_entry = ttk.Entry(search_frame, width=30)
        search_entry.pack(side=tk.LEFT, padx=5)

        columns = ('ID', 'Número de Hoja', 'Fecha', 'Cantidad Total', 'Total Venta')
        tree = ttk.Treeview(main_frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=180)
        
        tree.pack(fill=tk.BOTH, expand=True, pady=10)
        

        

        def export_to_excel():
            """Exporta el historial de ventas a un archivo Excel con dos reportes."""
            try:
                # Obtener fechas seleccionadas
                fecha_inicio = start_date.get_date()
                fecha_fin = end_date.get_date()
                
                # Abrir diálogo para seleccionar ubicación de guardado
                filename = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    initialfile=f"reporte_ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    title="Guardar reporte de ventas"
                )
                
                if not filename:  # Si el usuario cancela el diálogo
                    return
                    
                # Consulta para el reporte detallado con filtro de fechas
                self.manager.cursor.execute('''
                    SELECT 
                        h.numero_hoja,
                        p.nombre as nombre_producto,
                        h.fecha,
                        d.cantidad
                    FROM historial_ventas h
                    JOIN detalle_ventas d ON h.id = d.venta_id
                    JOIN productos p ON d.codigo_producto = p.codigo
                    WHERE DATE(h.fecha) BETWEEN DATE(?) AND DATE(?)
                    ORDER BY h.fecha DESC, h.numero_hoja
                ''', (fecha_inicio, fecha_fin))
                data = self.manager.cursor.fetchall()
                
                if not data:
                    messagebox.showwarning("Advertencia", "No hay datos para exportar en el rango de fechas seleccionado")
                    return
                
                wb = openpyxl.Workbook()
                
                # Configuración de estilos
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                total_font = Font(bold=True)
                alignment = Alignment(horizontal="center")
                
                # Reporte Detallado
                ws_detailed = wb.active
                ws_detailed.title = "Reporte Detallado"
                
                # Encabezados simplificados para reporte detallado
                detailed_columns = [
                    'Número de Hoja', 
                    'Producto', 
                    'Fecha', 
                    'Cantidad'
                ]
                
                for col_num, column_title in enumerate(detailed_columns, 1):
                    cell = ws_detailed.cell(row=1, column=col_num, value=column_title)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    ws_detailed.column_dimensions[cell.column_letter].width = max(len(column_title) + 2, 20)
                
                # Datos para reporte detallado
                for row_num, row_data in enumerate(data, 2):
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws_detailed.cell(row=row_num, column=col_num, value=value)
                        cell.alignment = alignment
                
                # Reporte de Ganancias
                ws_profits = wb.create_sheet(title="Reporte de Ganancias")
                
                # Encabezados para reporte de ganancias
                profit_columns = ['Código Producto', 'Nombre Producto', 'Cantidad Vendida', 
                                'Precio Costo', 'Precio Venta', 'Ganancia Unitaria', 'Ganancia Total']
                
                for col_num, column_title in enumerate(profit_columns, 1):
                    cell = ws_profits.cell(row=1, column=col_num, value=column_title)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    ws_profits.column_dimensions[cell.column_letter].width = max(len(column_title) + 2, 15)
                
                # Obtener datos de ganancias con filtro de fechas
                self.manager.cursor.execute('''
                    SELECT 
                        d.codigo_producto,
                        p.nombre,
                        SUM(d.cantidad) as cantidad_total,
                        p.precio_costo,
                        d.precio_unitario,
                        (d.precio_unitario - p.precio_costo) as ganancia_unitaria,
                        SUM(d.cantidad * (d.precio_unitario - p.precio_costo)) as ganancia_total
                    FROM detalle_ventas d
                    JOIN productos p ON d.codigo_producto = p.codigo
                    JOIN historial_ventas h ON d.venta_id = h.id
                    WHERE DATE(h.fecha) BETWEEN DATE(?) AND DATE(?)
                    GROUP BY d.codigo_producto, d.precio_unitario
                    ORDER BY ganancia_total DESC
                ''', (fecha_inicio, fecha_fin))
                
                profit_data = self.manager.cursor.fetchall()
                total_ganancia = 0
                total_costo = 0
                total_venta = 0
                
                # Llenar datos de ganancias
                for row_num, row_data in enumerate(profit_data, 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        cell = ws_profits.cell(row=row_num, column=col_num, value=cell_value)
                        cell.alignment = alignment
                        if col_num in [4, 5, 6, 7]:  # Formatear columnas de precios y ganancias
                            cell.number_format = 'Q#,##0.00'
                    
                    # Calcular totales
                    cantidad = row_data[2]
                    precio_costo = row_data[3]
                    precio_venta = row_data[4]
                    ganancia = row_data[6]
                    
                    total_costo += cantidad * precio_costo
                    total_venta += cantidad * precio_venta
                    total_ganancia += ganancia
                
                # Agregar totales al reporte de ganancias
                row_num = len(profit_data) + 3
                ws_profits.cell(row=row_num, column=1, value="TOTALES").font = total_font
                ws_profits.cell(row=row_num, column=4, value=total_costo).font = total_font
                ws_profits.cell(row=row_num, column=5, value=total_venta).font = total_font
                ws_profits.cell(row=row_num, column=7, value=total_ganancia).font = total_font
                
                ws_profits.cell(row=row_num, column=4).number_format = 'Q#,##0.00'
                ws_profits.cell(row=row_num, column=5).number_format = 'Q#,##0.00'
                ws_profits.cell(row=row_num, column=7).number_format = 'Q#,##0.00'
                
                wb.save(filename)
                messagebox.showinfo("Éxito", f"Reporte exportado como {filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Error al exportar a Excel: {str(e)}")

        def load_history_data(search_term=None):
            for item in tree.get_children():
                tree.delete(item)

            try:
                query = '''
                    SELECT h.id, h.numero_hoja, h.fecha, 
                           SUM(d.cantidad) AS total_cantidad,
                           h.total_venta
                    FROM historial_ventas h
                    JOIN detalle_ventas d ON h.id = d.venta_id
                    '''
                if search_term:
                    query += " WHERE h.numero_hoja LIKE ? "
                    self.manager.cursor.execute(query + " GROUP BY h.id ORDER BY h.fecha DESC", (f'%{search_term}%',))
                else:
                    self.manager.cursor.execute(query + " GROUP BY h.id ORDER BY h.fecha DESC")

                for row in self.manager.cursor.fetchall():
                    formatted_row = list(row)
                    formatted_row[4] = f"Q{formatted_row[4]:.2f}"  # Formatear total_venta
                    tree.insert('', tk.END, values=formatted_row)
            except Exception as e:
                messagebox.showerror("Error", f"Error al cargar el historial: {str(e)}")

        def view_details():
            """Muestra los detalles de una venta seleccionada."""
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Advertencia", "Seleccione una venta para ver los detalles")
                return

            venta_id = tree.item(selected[0])['values'][0]

            details_dialog = tk.Toplevel(dialog)
            details_dialog.title("Detalles de Venta")
            details_dialog.geometry("800x400")
            details_dialog.transient(dialog)
            details_dialog.grab_set()

            details_frame = ttk.Frame(details_dialog, padding="10")
            details_frame.pack(fill=tk.BOTH, expand=True)

            detail_columns = ('Código Producto', 'Cantidad', 'Precio Unitario', 'Subtotal')
            details_tree = ttk.Treeview(details_frame, columns=detail_columns, show='headings', height=10)

            for col in detail_columns:
                details_tree.heading(col, text=col)
                details_tree.column(col, width=150)

            details_tree.pack(fill=tk.BOTH, expand=True, pady=10)

            try:
                self.manager.cursor.execute('''
                    SELECT codigo_producto, cantidad, precio_unitario, subtotal
                    FROM detalle_ventas
                    WHERE venta_id = ?
                ''', (venta_id,))

                for row in self.manager.cursor.fetchall():
                    formatted_row = list(row)
                    formatted_row[2] = f"Q{formatted_row[2]:.2f}"
                    formatted_row[3] = f"Q{formatted_row[3]:.2f}"
                    details_tree.insert('', tk.END, values=formatted_row)
            except Exception as e:
                messagebox.showerror("Error", f"Error al cargar los detalles: {str(e)}")

        def delete_history():
            """Elimina todo el historial de ventas."""
            if not self.validate_admin_password():
                return
            if messagebox.askyesno("Confirmar", "¿Está seguro de eliminar todo el historial de ventas?"):
                try:
                    self.manager.cursor.execute("DELETE FROM detalle_ventas")
                    self.manager.cursor.execute("DELETE FROM historial_ventas")
                    self.manager.connection.commit()
                    load_history_data()
                    messagebox.showinfo("Éxito", "Historial eliminado correctamente")
                except Exception as e:
                    messagebox.showerror("Error", f"Error al eliminar el historial: {str(e)}")
        def delete_selected_sale():
            """Elimina la venta seleccionada y restaura los productos al inventario."""
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Advertencia", "Seleccione una venta para eliminar")
                return

            if not self.validate_admin_password():
                return

            if not messagebox.askyesno("Confirmar", "¿Está seguro de eliminar la venta seleccionada? Los productos serán devueltos al inventario."):
                return

            venta_id = tree.item(selected[0])['values'][0]

            try:
                # Primero obtener los detalles de la venta para restaurar el inventario
                self.manager.cursor.execute('''
                    SELECT codigo_producto, cantidad
                    FROM detalle_ventas
                    WHERE venta_id = ?
                ''', (venta_id,))
                
                detalles_venta = self.manager.cursor.fetchall()

                # Restaurar el inventario
                for codigo_producto, cantidad in detalles_venta:
                    self.manager.cursor.execute('''
                        UPDATE productos 
                        SET cantidad = cantidad + ?
                        WHERE codigo = ?
                    ''', (cantidad, codigo_producto))

                # Eliminar los registros de la venta
                self.manager.cursor.execute("DELETE FROM detalle_ventas WHERE venta_id = ?", (venta_id,))
                self.manager.cursor.execute("DELETE FROM historial_ventas WHERE id = ?", (venta_id,))
                
                self.manager.connection.commit()
                load_history_data()
                messagebox.showinfo("Éxito", "Venta eliminada y productos restaurados correctamente")
            except Exception as e:
                self.manager.connection.rollback()
                messagebox.showerror("Error", f"Error al eliminar la venta: {str(e)}")
        def search_history(*args):
            load_history_data(search_entry.get().strip())

        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=5)
    
        ttk.Button(button_frame, text="Buscar", command=search_history).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Ver Detalles", command=view_details).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Exportar a Excel", command=export_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Eliminar Seleccionado", command=delete_selected_sale).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Eliminar Historial", command=delete_history).pack(side=tk.LEFT, padx=5)
        
        search_entry.bind('<Return>', search_history)
    
        # Cargar datos iniciales
        load_history_data()
    def view_return_history(self):
     """Muestra el historial de devoluciones en una ventana nueva."""
     if not self.manager.current_db:
        messagebox.showerror("Error", "No hay una base de datos seleccionada")
        return

     dialog = tk.Toplevel(self.root)
     dialog.title("Historial de Devoluciones")
     dialog.geometry("1300x600")  # Aumentado para acomodar la nueva columna
     dialog.transient(self.root)
     dialog.grab_set()

     main_frame = ttk.Frame(dialog, padding="10")
     main_frame.pack(fill=tk.BOTH, expand=True)

    # Frame de búsqueda
     search_frame = ttk.Frame(main_frame)
     search_frame.pack(fill=tk.X, pady=(0, 10))
    
     ttk.Label(search_frame, text="Buscar por número de envío:").pack(side=tk.LEFT, padx=5)
     search_entry = ttk.Entry(search_frame, width=30)
     search_entry.pack(side=tk.LEFT, padx=5)

    # Tabla para el historial de devoluciones
     columns = ('ID', 'Número de Envío', 'Fecha de Envío', 'Fecha de Devolución', 'Total Productos')
     tree = ttk.Treeview(main_frame, columns=columns, show='headings', height=15)

    # Ajustar anchos de columna
     column_widths = {
        'ID': 80,
        'Número de Envío': 200,
        'Fecha de Envío': 200,
        'Fecha de Devolución': 200,
        'Total Productos': 150
    }

     for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=column_widths.get(col, 200))

     tree.pack(fill=tk.BOTH, expand=True, pady=10)

     
     def view_details():
        """Muestra los detalles de una devolución seleccionada"""
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Advertencia", "Seleccione una devolución para ver los detalles")
            return

        devolucion_id = tree.item(selected[0])['values'][0]

        details_dialog = tk.Toplevel(dialog)
        details_dialog.title("Detalles de Devolución")
        details_dialog.geometry("1000x400")
        details_dialog.transient(dialog)
        details_dialog.grab_set()

        details_frame = ttk.Frame(details_dialog, padding="10")
        details_frame.pack(fill=tk.BOTH, expand=True)

        # Tabla de detalles con columnas adicionales
        detail_columns = ('Código Producto', 'Cantidad Antes', 'Cantidad Devuelta', 'Cantidad Después', 
                         'Precio Unitario', 'Subtotal')
        details_tree = ttk.Treeview(details_frame, columns=detail_columns, show='headings', height=10)
        
        # Configurar columnas con anchos apropiados
        column_widths = {
            'Código Producto': 120,
            'Cantidad Antes': 120,
            'Cantidad Devuelta': 120,
            'Cantidad Después': 120,
            'Precio Unitario': 120,
            'Subtotal': 120
        }
        
        for col in detail_columns:
            details_tree.heading(col, text=col)
            details_tree.column(col, width=column_widths.get(col, 120))

        details_tree.pack(fill=tk.BOTH, expand=True, pady=10)

        # Cargar detalles de la devolución seleccionada
        try:
            self.manager.cursor.execute('''
                SELECT 
                    d.codigo_producto,
                    d.cantidad + p.cantidad as cantidad_antes,
                    d.cantidad as cantidad_devuelta,
                    p.cantidad as cantidad_despues,
                    d.precio_unitario,
                    d.subtotal
                FROM detalle_devolucion d
                JOIN productos p ON d.codigo_producto = p.codigo
                WHERE d.devolucion_id = ?
            ''', (devolucion_id,))
            
            for detail_row in self.manager.cursor.fetchall():
                # Formatear los valores monetarios y crear lista para la fila
                formatted_detail = [
                    detail_row[0],                    # Código Producto
                    detail_row[1],                    # Cantidad Antes
                    detail_row[2],                    # Cantidad Devuelta
                    detail_row[3],                    # Cantidad Después
                    f"Q{float(detail_row[4]):.2f}",  # Precio Unitario
                    f"Q{float(detail_row[5]):.2f}"   # Subtotal
                ]
                details_tree.insert('', tk.END, values=formatted_detail)
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar los detalles: {str(e)}")

     def load_history_data(search_term=None):
        """Carga los datos del historial con opción de filtrado"""
        for item in tree.get_children():
            tree.delete(item)
            
        try:
            if search_term:
                query = '''
                    SELECT 
                        h.id, 
                        h.numero_envio,
                        h.fecha_envio,
                        h.fecha_devolucion,
                        SUM(d.cantidad) as total_cantidad
                    FROM historial_devoluciones h
                    JOIN detalle_devolucion d ON h.id = d.devolucion_id
                    WHERE h.numero_envio LIKE ?
                    GROUP BY h.id, h.numero_envio, h.fecha_envio, h.fecha_devolucion
                    ORDER BY h.fecha_devolucion DESC
                '''
                self.manager.cursor.execute(query, (f'%{search_term}%',))
            else:
                query = '''
                    SELECT 
                        h.id, 
                        h.numero_envio,
                        h.fecha_envio,
                        h.fecha_devolucion,
                        SUM(d.cantidad) as total_cantidad
                    FROM historial_devoluciones h
                    JOIN detalle_devolucion d ON h.id = d.devolucion_id
                    GROUP BY h.id, h.numero_envio, h.fecha_envio, h.fecha_devolucion
                    ORDER BY h.fecha_devolucion DESC
                '''
                self.manager.cursor.execute(query)
                
            for row in self.manager.cursor.fetchall():
                tree.insert('', tk.END, values=row)
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar el historial: {str(e)}")

     def delete_selected():
        """Elimina la devolución seleccionada y restaura los productos"""
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Advertencia", "Seleccione una devolución para eliminar")
            return

        if not self.validate_admin_password():
            return

        if not messagebox.askyesno("Confirmar", "¿Está seguro que desea eliminar esta devolución? Los productos serán devueltos al inventario."):
            return

        devolucion_id = tree.item(selected[0])['values'][0]

        try:
            # Primero obtenemos los detalles de la devolución
            self.manager.cursor.execute('''
                SELECT codigo_producto, cantidad
                FROM detalle_devolucion
                WHERE devolucion_id = ?
            ''', (devolucion_id,))
            
            detalles = self.manager.cursor.fetchall()
            
            # Actualizamos el inventario sumando las cantidades devueltas
            for codigo_producto, cantidad in detalles:
                self.manager.cursor.execute('''
                    UPDATE productos 
                    SET cantidad = cantidad + ?
                    WHERE codigo = ?
                ''', (cantidad, codigo_producto))
            
            # Eliminamos los registros de la devolución
            self.manager.cursor.execute("DELETE FROM detalle_devolucion WHERE devolucion_id = ?", (devolucion_id,))
            self.manager.cursor.execute("DELETE FROM historial_devoluciones WHERE id = ?", (devolucion_id,))
            
            self.manager.connection.commit()
            load_history_data()
            messagebox.showinfo("Éxito", "Devolución eliminada y productos restaurados correctamente")
        except Exception as e:
            self.manager.connection.rollback()
            messagebox.showerror("Error", f"Error al eliminar la devolución: {str(e)}")
     def search_shipment():
        """Busca devoluciones por número de envío"""
        search_term = search_entry.get().strip()
        load_history_data(search_term)

     def clear_history():
        """Elimina todo el historial de devoluciones"""
        if not self.validate_admin_password():
            return
        if messagebox.askyesno("Confirmar", "¿Está seguro que desea eliminar todo el historial de devoluciones?"):
            try:
                self.manager.cursor.execute("DELETE FROM detalle_devolucion")
                self.manager.cursor.execute("DELETE FROM historial_devoluciones")
                self.manager.connection.commit()
                load_history_data()
                messagebox.showinfo("Éxito", "Historial eliminado correctamente")
            except Exception as e:
                messagebox.showerror("Error", f"Error al eliminar el historial: {str(e)}")

     

     button_frame = ttk.Frame(main_frame)
     button_frame.pack(fill=tk.X, pady=10)

     ttk.Button(search_frame, text="Buscar", command=lambda: load_history_data(search_entry.get().strip())).pack(side=tk.LEFT, padx=5)
     ttk.Button(button_frame, text="Ver Detalles", command=view_details).pack(side=tk.LEFT, padx=5)
     ttk.Button(button_frame, text="Exportar a Excel", command=lambda: self.export_to_report()).pack(side=tk.LEFT, padx=5)
     ttk.Button(button_frame, text="Eliminar Seleccionada", command=delete_selected).pack(side=tk.LEFT, padx=5)
     ttk.Button(button_frame, text="Eliminar Historial", command=clear_history).pack(side=tk.LEFT, padx=5)

    # Cargar datos iniciales
     load_history_data()

    # Vincular la tecla Enter al botón de búsqueda
     search_entry.bind('<Return>', lambda e: load_history_data(search_entry.get().strip()))

     
    def view_purchase_history(self):
        """Muestra el historial de compras en una ventana nueva"""
        if not self.manager.current_db:
            messagebox.showerror("Error", "No hay una base de datos seleccionada")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Historial de Compras")
        dialog.geometry("1000x600")
        dialog.transient(self.root)
        dialog.grab_set()

        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Frame para búsqueda
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=5)

        ttk.Label(search_frame, text="Buscar por número de envío:").pack(side=tk.LEFT, padx=5)
        search_entry = ttk.Entry(search_frame, width=30)
        search_entry.pack(side=tk.LEFT, padx=5)

        # Tabla para el historial de compras
        columns = ('ID', 'Número de Envío', 'Fecha', 'Cantidad Total de Productos', 'Total Compra')
        tree = ttk.Treeview(main_frame, columns=columns, show='headings', height=15)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=180)

        tree.pack(fill=tk.BOTH, expand=True, pady=10)

        def view_details():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Advertencia", "Seleccione una compra para ver los detalles")
                return

            compra_id = tree.item(selected[0])['values'][0]

            details_dialog = tk.Toplevel(dialog)
            details_dialog.title("Detalles de Compra")
            details_dialog.geometry("800x400")
            details_dialog.transient(dialog)
            details_dialog.grab_set()

            details_frame = ttk.Frame(details_dialog, padding="10")
            details_frame.pack(fill=tk.BOTH, expand=True)

            # Tabla de detalles
            detail_columns = ('Código Producto', 'Cantidad Antes', 'Cantidad Devuelta', 'Cantidad Después', 
                            'Precio Unitario', 'Subtotal')
            details_tree = ttk.Treeview(details_frame, columns=detail_columns, show='headings', height=10)

            column_widths = {
                'Código Producto': 120,
                'Cantidad Antes': 120,
                'Cantidad Devuelta': 120,
                'Cantidad Después': 120,
                'Precio Unitario': 120,
                'Subtotal': 120
            }
            for col in detail_columns:
                details_tree.heading(col, text=col)
                details_tree.column(col, width=column_widths.get(col, 120))

            details_tree.pack(fill=tk.BOTH, expand=True, pady=10)

            # Cargar detalles de la compra seleccionada
            try:
                self.manager.cursor.execute('''
                    SELECT 
                        d.codigo_producto,
                        d.cantidad + p.cantidad as cantidad_antes,
                        d.cantidad as cantidad_devuelta,
                        p.cantidad as cantidad_despues,
                        d.precio_unitario,
                        d.subtotal
                    FROM detalle_compras d
                    JOIN productos p ON d.codigo_producto = p.codigo
                    WHERE d.compra_id = ?
                ''', (compra_id,))
                for detail_row in self.manager.cursor.fetchall():
                    formatted_detail = [
                        detail_row[0],                    # Código Producto
                        detail_row[1],                    # Cantidad Antes
                        detail_row[2],                    # Cantidad Devuelta
                        detail_row[3],                    # Cantidad Después
                        f"Q{float(detail_row[4]):.2f}",  # Precio Unitario
                        f"Q{float(detail_row[5]):.2f}"   # Subtotal
                    ]
                    details_tree.insert('', tk.END, values=formatted_detail)
            except Exception as e:
                messagebox.showerror("Error", f"Error al cargar los detalles: {str(e)}")

        def load_history_data(search_term=None):
            for item in tree.get_children():
                tree.delete(item)

            try:
                if search_term:
                    query = '''
                        SELECT h.id, h.numero_envio, h.fecha, 
                               SUM(d.cantidad) AS total_cantidad,
                               h.total_compra
                        FROM historial_compras h
                        JOIN detalle_compras d ON h.id = d.compra_id
                        WHERE h.numero_envio LIKE ?
                        GROUP BY h.id
                        ORDER BY h.fecha DESC
                    '''
                    self.manager.cursor.execute(query, (f'%{search_term}%',))
                else:
                    query = '''
                        SELECT h.id, h.numero_envio, h.fecha, 
                               SUM(d.cantidad) AS total_cantidad,
                               h.total_compra
                        FROM historial_compras h
                        JOIN detalle_compras d ON h.id = d.compra_id
                        GROUP BY h.id
                        ORDER BY h.fecha DESC
                    '''
                    self.manager.cursor.execute(query)

                for row in self.manager.cursor.fetchall():
                    formatted_row = list(row)
                    formatted_row[4] = f"Q{formatted_row[4]:.2f}"  # Format total_compra
                    tree.insert('', tk.END, values=formatted_row)
            except Exception as e:
                messagebox.showerror("Error", f"Error al cargar el historial: {str(e)}")

        def search_history(*args):
            load_history_data(search_entry.get().strip())
        def delete_selected():
            """Elimina la compra seleccionada y resta los productos del inventario"""
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Advertencia", "Seleccione una compra para eliminar")
                return

            if not self.validate_admin_password():
                return

            if not messagebox.askyesno("Confirmar", "¿Está seguro que desea eliminar esta compra? Los productos serán restados del inventario."):
                return

            compra_id = tree.item(selected[0])['values'][0]

            try:
                # Primero obtenemos los detalles de la compra
                self.manager.cursor.execute('''
                    SELECT codigo_producto, cantidad
                    FROM detalle_compras
                    WHERE compra_id = ?
                ''', (compra_id,))
                
                detalles = self.manager.cursor.fetchall()
                
                # Actualizamos el inventario restando las cantidades compradas
                for codigo_producto, cantidad in detalles:
                    self.manager.cursor.execute('''
                        UPDATE productos 
                        SET cantidad = cantidad - ?
                        WHERE codigo = ?
                    ''', (cantidad, codigo_producto))
                    
                    # Verificar que no queden cantidades negativas
                    self.manager.cursor.execute('''
                        SELECT cantidad 
                        FROM productos 
                        WHERE codigo = ? AND cantidad < 0
                    ''', (codigo_producto,))
                    
                    if self.manager.cursor.fetchone():
                        raise Exception("La eliminación dejaría cantidades negativas en el inventario")
                
                # Eliminamos los registros de la compra
                self.manager.cursor.execute("DELETE FROM detalle_compras WHERE compra_id = ?", (compra_id,))
                self.manager.cursor.execute("DELETE FROM historial_compras WHERE id = ?", (compra_id,))
                
                self.manager.connection.commit()
                load_history_data()
                messagebox.showinfo("Éxito", "Compra eliminada y productos actualizados correctamente")
            except Exception as e:
                self.manager.connection.rollback()
                messagebox.showerror("Error", f"Error al eliminar la compra: {str(e)}")
        def clear_history():
            if not self.validate_admin_password():
                return
            if messagebox.askyesno("Confirmar", "¿Está seguro de borrar todo el historial de compras?"):
                try:
                    self.manager.cursor.execute("DELETE FROM detalle_compras")
                    self.manager.cursor.execute("DELETE FROM historial_compras")
                    self.manager.connection.commit()
                    load_history_data()
                    messagebox.showinfo("Éxito", "Historial borrado correctamente")
                except Exception as e:
                    messagebox.showerror("Error", f"Error al borrar el historial: {str(e)}")

        def generate_excel_report():
            """Genera un reporte en Excel del historial de compras"""
            try:
                # Importar openpyxl solo cuando se necesite
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment
                
                # Crear un nuevo libro de Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Reporte de Compras"
                
                # Definir encabezados
                headers = ['Número de Envío', 'Nombre del Producto', 'Fecha', 'Cantidad']
                
                # Dar formato a los encabezados
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # Obtener datos de la base de datos
                query = '''
                    SELECT 
                        h.numero_envio,
                        p.nombre,
                        h.fecha,
                        d.cantidad
                    FROM historial_compras h
                    JOIN detalle_compras d ON h.id = d.compra_id
                    JOIN productos p ON d.codigo_producto = p.codigo
                    ORDER BY h.fecha DESC, h.numero_envio
                '''
                
                self.manager.cursor.execute(query)
                data = self.manager.cursor.fetchall()
                
                # Llenar datos
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        # Formatear fecha si es necesario
                        if col_idx == 3 and isinstance(value, str):
                            try:
                                fecha = datetime.strptime(value, '%Y-%m-%d')
                                value = fecha.strftime('%d/%m/%Y')
                            except:
                                pass
                        cell.value = value
                        cell.alignment = Alignment(horizontal='center')
                
                # Ajustar el ancho de las columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Solicitar ubicación para guardar el archivo
                filename = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    title="Guardar reporte de compras"
                )
                
                if filename:
                    wb.save(filename)
                    messagebox.showinfo("Éxito", "Reporte generado correctamente")
                    
            except ImportError:
                messagebox.showerror("Error", "Por favor instale openpyxl: pip install openpyxl")
            except Exception as e:
                messagebox.showerror("Error", f"Error al generar el reporte: {str(e)}")

        # Botones
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=5)

        ttk.Button(button_frame, text="Buscar", command=search_history).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Ver Detalles", command=view_details).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Eliminar Seleccionada", command=delete_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Borrar Historial", command=clear_history).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Generar Reporte Excel", command=generate_excel_report).pack(side=tk.LEFT, padx=5)

        # Bind search entry to Return key
        search_entry.bind('<Return>', search_history)

        # Initial load of data
        load_history_data()




    
    def validate_admin_password(self):
        """Valida la contraseña de administrador"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Validación de Administrador")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Ingrese la contraseña de administrador:").pack(pady=5)
        password_entry = ttk.Entry(frame, show="•")
        password_entry.pack(pady=5)
        
        result = [False]  # Usar lista para poder modificar desde la función interna
        
        def check_password():
            if password_entry.get() == self.ADMIN_PASSWORD:
                result[0] = True
                dialog.destroy()
            else:
                messagebox.showerror("Error", "Contraseña incorrecta")
                dialog.lift()
        
        ttk.Button(frame, text="Validar", command=check_password).pack(pady=10)
        
        dialog.wait_window()  # Esperar hasta que se cierre la ventana
        return result[0]
    
    def setup_keyboard_navigation(entries, submit_action):
     """
    Configura la navegación por teclado para un conjunto de campos
    
    Args:
        entries: Lista de campos Entry en orden de tabulación
        submit_action: Función a ejecutar cuando se presiona Enter en el último campo
     """
     def focus_next_widget(event):
        current = event.widget
        try:
            next_index = entries.index(current) + 1
            if next_index < len(entries):
                entries[next_index].focus_set()
            else:
                submit_action()
        except ValueError:
            pass
        return "break"  # Previene el comportamiento por defecto

    # Configurar eventos para cada campo
     for entry in entries:
        entry.bind('<Return>', focus_next_widget)
        entry.bind('<Tab>', focus_next_widget)    
    def make_return(self):
     """Permite realizar una devolución con múltiples productos y registra el historial."""
     if not self.manager.current_db:
        messagebox.showerror("Error", "No hay una base de datos seleccionada")
        return

    # Crear tablas si no existen
     self.manager.cursor.execute('''
        CREATE TABLE IF NOT EXISTS historial_devoluciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_envio TEXT NOT NULL,
            fecha_envio DATE NOT NULL,
            fecha_devolucion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            total_devoluciones REAL NOT NULL
        )
    ''')

     self.manager.cursor.execute('''
        CREATE TABLE IF NOT EXISTS detalle_devolucion (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            devolucion_id INTEGER,
            codigo_producto TEXT NOT NULL,
            cantidad INTEGER NOT NULL,
            precio_unitario REAL NOT NULL,
            subtotal REAL NOT NULL,
            FOREIGN KEY (devolucion_id) REFERENCES historial_devoluciones(id)
        )
    ''')
    
     self.manager.connection.commit()

     dialog = tk.Toplevel(self.root)
     dialog.title("Realizar Devolución")
     dialog.geometry("1000x800")
     dialog.transient(self.root)
     dialog.grab_set()

     main_frame = ttk.Frame(dialog, padding="20")
     main_frame.pack(fill=tk.BOTH, expand=True)
    
    # Frame para información de envío
     shipment_frame = ttk.LabelFrame(main_frame, text="Información de Envío", padding="10")
     shipment_frame.pack(fill=tk.X, pady=(0, 10))
    
     ttk.Label(shipment_frame, text="Número de Envío:").pack(side=tk.LEFT, padx=5)
     shipment_number = ttk.Entry(shipment_frame, width=30)
     shipment_number.pack(side=tk.LEFT, padx=5)
    
     cal_frame = ttk.Frame(shipment_frame)
     cal_frame.pack(side=tk.RIGHT, padx=5)
     ttk.Label(cal_frame, text="Fecha de Envío Original:").pack()
     cal = Calendar(cal_frame, selectmode='day', date_pattern='yyyy-mm-dd',
                   locale='es_ES', weekenddays=[6,7],
                   weekendbackground='white', weekendforeground='black')
     cal.pack(pady=5)

    # Frame para productos
     products_frame = ttk.LabelFrame(main_frame, text="Productos", padding="10")
     products_frame.pack(fill=tk.BOTH, expand=True)
    
     columns = ('Código', 'Nombre', 'Cantidad', 'Precio Unitario', 'Subtotal')
     tree = ttk.Treeview(products_frame, columns=columns, show='headings', height=10)
    
     for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=100)
    
     tree.pack(fill=tk.BOTH, expand=True, pady=5)
    
     entry_frame = ttk.Frame(products_frame)
     entry_frame.pack(fill=tk.X, pady=5)

     ttk.Label(entry_frame, text="Código:").grid(row=0, column=0, padx=5)
     code_entry = ttk.Entry(entry_frame, width=20)
     code_entry.grid(row=0, column=1, padx=5)

     ttk.Label(entry_frame, text="Cantidad:").grid(row=0, column=2, padx=5)
     quantity_entry = ttk.Entry(entry_frame, width=10)
     quantity_entry.grid(row=0, column=3, padx=5)
    
     product_info = {
        'name': tk.StringVar(),
        'price': tk.StringVar(),
        'stock': tk.StringVar()
    }

     info_frame = ttk.LabelFrame(products_frame, text="Información del Producto", padding="5")
     info_frame.pack(fill=tk.X, pady=5)

     ttk.Label(info_frame, textvariable=product_info['name']).pack()
     ttk.Label(info_frame, textvariable=product_info['price']).pack()
     ttk.Label(info_frame, textvariable=product_info['stock']).pack()
    
     def search_product(*args):
        code = code_entry.get().strip()
        if code:
            try:
                self.manager.cursor.execute('''
                    SELECT nombre, precio_costo, cantidad
                    FROM productos
                    WHERE codigo = ?
                ''', (code,))
                
                result = self.manager.cursor.fetchone()
                if result:
                    product_info['name'].set(f"Nombre: {result[0]}")
                    product_info['price'].set(f"Precio de costo: Q{result[1]:.2f}")
                    product_info['stock'].set(f"Stock actual: {result[2]}")
                else:
                    product_info['name'].set("Producto no encontrado")
                    product_info['price'].set("")
                    product_info['stock'].set("")
            except Exception as e:
                messagebox.showerror("Error", f"Error al buscar el producto: {str(e)}")
    
     def add_to_list():
        code = code_entry.get().strip()
        quantity = quantity_entry.get().strip()

        if not code or not quantity:
            messagebox.showwarning("Advertencia", "Por favor ingrese código y cantidad")
            return

        try:
            quantity = int(quantity)
            if quantity <= 0:
                messagebox.showerror("Error", "La cantidad debe ser mayor a 0")
                return

            self.manager.cursor.execute('''
                SELECT nombre, cantidad
                FROM productos
                WHERE codigo = ?
            ''', (code,))

            result = self.manager.cursor.fetchone()
            if result:
                subtotal = quantity * result[1]
                tree.insert('', tk.END, values=(
                    code,
                    result[0],
                    quantity,
                    f"Q{result[1]:.2f}",
                    f"Q{subtotal:.2f}"
                ))

                # Limpiar campos
                code_entry.delete(0, tk.END)
                quantity_entry.delete(0, tk.END)
                product_info['name'].set("")
                product_info['price'].set("")
                product_info['stock'].set("")
                code_entry.focus_set()
            else:
                messagebox.showerror("Error", "Producto no encontrado")

        except ValueError:
            messagebox.showerror("Error", "La cantidad debe ser un número entero")

      

     def process_return():
        if not shipment_number.get().strip():
            messagebox.showwarning("Advertencia", "Por favor ingrese un número de envío")
            return
        if not tree.get_children():
            messagebox.showwarning("Advertencia", "No hay productos en la lista")
            return

        try:
            total = 0
            for item in tree.get_children():
                values = tree.item(item)['values']
                subtotal = float(values[4].replace('Q', ''))
                total += subtotal

            # Obtener la fecha seleccionada del calendario
            fecha_envio = cal.get_date()

            # Insertar en historial_devoluciones con la fecha de envío
            self.manager.cursor.execute('''
                INSERT INTO historial_devoluciones 
                (numero_envio, fecha_envio, total_devoluciones)
                VALUES (?, ?, ?)
            ''', (shipment_number.get().strip(), fecha_envio, total))

            devolucion_id = self.manager.cursor.lastrowid

            # Insertar detalles y actualizar inventario
            for item in tree.get_children():
                values = tree.item(item)['values']
                codigo = values[0]
                cantidad = int(values[2])
                precio = float(values[3].replace('Q', ''))
                subtotal = float(values[4].replace('Q', ''))

                self.manager.cursor.execute('''
                    INSERT INTO detalle_devolucion 
                    (devolucion_id, codigo_producto, cantidad, precio_unitario, subtotal)
                    VALUES (?, ?, ?, ?, ?)
                ''', (devolucion_id, codigo, cantidad, precio, subtotal))

                self.manager.cursor.execute('''
                    UPDATE productos
                    SET cantidad = cantidad - ?
                    WHERE codigo = ?
                ''', (cantidad, codigo))

            self.manager.connection.commit()
            self.load_inventory_data()
            messagebox.showinfo("Éxito", "Devolución procesada correctamente")
            dialog.destroy()

        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar la devolución: {str(e)}")

      
    
    # Botones
     button_frame = ttk.Frame(products_frame)
     button_frame.pack(fill=tk.X, pady=10)
     ttk.Button(button_frame, text="Agregar Producto", command=add_to_list).pack(side=tk.LEFT, padx=5)
     ttk.Button(button_frame, text="Eliminar Seleccionado", command=lambda: tree.delete(tree.selection())).pack(side=tk.LEFT, padx=5)
     ttk.Button(button_frame, text="Procesar Devolución", command=process_return).pack(side=tk.RIGHT, padx=5)
    
     code_entry.bind('<Return>', lambda e: search_product())
     quantity_entry.bind('<Return>', lambda e: add_to_list())
     code_entry.focus_set()



    
    def make_purchase(self):
     """Muestra la ventana para realizar una compra con múltiples productos"""
     if not self.manager.current_db:
        messagebox.showerror("Error", "No hay una base de datos seleccionada")
        return
    

    # Crear tabla de historial de compras si no existe
     self.manager.cursor.execute('''
        CREATE TABLE IF NOT EXISTS historial_compras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_envio TEXT NOT NULL,
            fecha TIMESTAMP NOT NULL,
            total_compra REAL NOT NULL
        )
    ''')

     self.manager.cursor.execute('''
        CREATE TABLE IF NOT EXISTS detalle_compras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            compra_id INTEGER,
            codigo_producto TEXT NOT NULL,
            cantidad INTEGER NOT NULL,
            precio_unitario REAL NOT NULL,
            subtotal REAL NOT NULL,
            FOREIGN KEY (compra_id) REFERENCES historial_compras(id)
        )
    ''')
     self.manager.connection.commit()

     dialog = tk.Toplevel(self.root)
     dialog.title("Realizar Compra")
     dialog.geometry("1000x800")  # Increased height for calendar
     dialog.transient(self.root)
     dialog.grab_set()

     main_frame = ttk.Frame(dialog, padding="20")
     main_frame.pack(fill=tk.BOTH, expand=True)
    
    # Frame para número de envío y calendario
     shipment_frame = ttk.LabelFrame(main_frame, text="Información de Envío", padding="10")
     shipment_frame.pack(fill=tk.X, pady=(0, 10))

    # Número de envío
     ttk.Label(shipment_frame, text="Número de Envío:").pack(side=tk.LEFT, padx=5)
     shipment_number = ttk.Entry(shipment_frame, width=30)
     shipment_number.pack(side=tk.LEFT, padx=5)

    # Calendario
     cal_frame = ttk.Frame(shipment_frame)
     cal_frame.pack(side=tk.RIGHT, padx=5)
     ttk.Label(cal_frame, text="Fecha de Envío:").pack()
     cal = Calendar(cal_frame, selectmode='day', date_pattern='yyyy-mm-dd',
                   locale='es_ES',
                   weekenddays=[6,7],
                   weekendbackground='white',
                   weekendforeground='black')
     cal.pack(pady=5)

    # Frame para productos
     products_frame = ttk.LabelFrame(main_frame, text="Productos", padding="10")
     products_frame.pack(fill=tk.BOTH, expand=True)

    # Lista de productos en la compra
     columns = ('Código', 'Nombre', 'Cantidad', 'Precio Unitario', 'Subtotal')
     tree = ttk.Treeview(products_frame, columns=columns, show='headings', height=10)
    
     for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=100)

     tree.pack(fill=tk.BOTH, expand=True, pady=5)

    # Frame para entrada de productos
     entry_frame = ttk.Frame(products_frame)
     entry_frame.pack(fill=tk.X, pady=5)

     ttk.Label(entry_frame, text="Código:").grid(row=0, column=0, padx=5)
     code_entry = ttk.Entry(entry_frame, width=20)
     code_entry.grid(row=0, column=1, padx=5)

     ttk.Label(entry_frame, text="Cantidad:").grid(row=0, column=2, padx=5)
     quantity_entry = ttk.Entry(entry_frame, width=10)
     quantity_entry.grid(row=0, column=3, padx=5)

    # Variables para mostrar información del producto
     product_info = {
        'name': tk.StringVar(),
        'price': tk.StringVar(),
        'stock': tk.StringVar()
    }

     info_frame = ttk.LabelFrame(products_frame, text="Información del Producto", padding="5")
     info_frame.pack(fill=tk.X, pady=5)

     ttk.Label(info_frame, textvariable=product_info['name']).pack()
     ttk.Label(info_frame, textvariable=product_info['price']).pack()
     ttk.Label(info_frame, textvariable=product_info['stock']).pack()

     def search_product(*args):
        code = code_entry.get().strip()
        if code:
            try:
                self.manager.cursor.execute('''
                    SELECT nombre, precio_costo, cantidad
                    FROM productos
                    WHERE codigo = ?
                ''', (code,))
                
                result = self.manager.cursor.fetchone()
                if result:
                    product_info['name'].set(f"Nombre: {result[0]}")
                    product_info['price'].set(f"Precio de costo: Q{result[1]:.2f}")
                    product_info['stock'].set(f"Stock actual: {result[2]}")
                else:
                    product_info['name'].set("Producto no encontrado")
                    product_info['price'].set("")
                    product_info['stock'].set("")
            except Exception as e:
                messagebox.showerror("Error", f"Error al buscar el producto: {str(e)}")

     def add_to_list():
        code = code_entry.get().strip()
        quantity = quantity_entry.get().strip()

        if not code or not quantity:
            messagebox.showwarning("Advertencia", "Por favor ingrese código y cantidad")
            return

        try:
            quantity = int(quantity)
            if quantity <= 0:
                messagebox.showerror("Error", "La cantidad debe ser mayor a 0")
                return

            self.manager.cursor.execute('''
                SELECT nombre, precio_costo
                FROM productos
                WHERE codigo = ?
            ''', (code,))

            result = self.manager.cursor.fetchone()
            if result:
                subtotal = quantity * result[1]
                tree.insert('', tk.END, values=(
                    code,
                    result[0],
                    quantity,
                    f"Q{result[1]:.2f}",
                    f"Q{subtotal:.2f}"
                ))
                
                # Limpiar campos
                code_entry.delete(0, tk.END)
                quantity_entry.delete(0, tk.END)
                product_info['name'].set("")
                product_info['price'].set("")
                product_info['stock'].set("")
                code_entry.focus_set()
            else:
                messagebox.showerror("Error", "Producto no encontrado")

        except ValueError:
            messagebox.showerror("Error", "La cantidad debe ser un número entero")

     def process_purchase():
        if not shipment_number.get().strip():
            messagebox.showwarning("Advertencia", "Por favor ingrese un número de envío")
            return

        if not tree.get_children():
            messagebox.showwarning("Advertencia", "No hay productos en la lista")
            return

        try:
            # Get selected date from calendar
            selected_date = cal.get_date()

            # Calcular total de la compra
            total = 0
            for item in tree.get_children():
                values = tree.item(item)['values']
                subtotal = float(values[4].replace('Q', ''))
                total += subtotal

            # Insertar en historial_compras con la fecha seleccionada
            self.manager.cursor.execute('''
                INSERT INTO historial_compras (numero_envio, fecha, total_compra)
                VALUES (?, ?, ?)
            ''', (shipment_number.get().strip(), selected_date, total))
            
            compra_id = self.manager.cursor.lastrowid

            # Insertar detalles y actualizar inventario
            for item in tree.get_children():
                values = tree.item(item)['values']
                codigo = values[0]
                cantidad = int(values[2])
                precio = float(values[3].replace('Q', ''))
                subtotal = float(values[4].replace('Q', ''))

                # Insertar en detalle_compras
                self.manager.cursor.execute('''
                    INSERT INTO detalle_compras 
                    (compra_id, codigo_producto, cantidad, precio_unitario, subtotal)
                    VALUES (?, ?, ?, ?, ?)
                ''', (compra_id, codigo, cantidad, precio, subtotal))

                # Actualizar inventario
                self.manager.cursor.execute('''
                    UPDATE productos
                    SET cantidad = cantidad + ?
                    WHERE codigo = ?
                ''', (cantidad, codigo))

            self.manager.connection.commit()
            self.load_inventory_data()
            messagebox.showinfo("Éxito", f"Compra procesada correctamente\nTotal: Q{total:.2f}")
            dialog.destroy()

        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar la compra: {str(e)}")

    # Botones
     button_frame = ttk.Frame(products_frame)
     button_frame.pack(fill=tk.X, pady=10)

     ttk.Button(button_frame, text="Agregar Producto", command=add_to_list).pack(side=tk.LEFT, padx=5)
     ttk.Button(button_frame, text="Eliminar Seleccionado", 
               command=lambda: tree.delete(tree.selection())).pack(side=tk.LEFT, padx=5)
     ttk.Button(button_frame, text="Procesar Compra", 
               command=process_purchase).pack(side=tk.RIGHT, padx=5)

    # Configurar eventos
     code_entry.bind('<Return>', lambda e: search_product())
     quantity_entry.bind('<Return>', lambda e: add_to_list())

     code_entry.focus_set()
    
    def make_sale(self):
     """Permite realizar una venta y registra el historial de ventas."""
     if not self.manager.current_db:
        messagebox.showerror("Error", "No hay una base de datos seleccionada")
        return

     try:
        locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
     except:
        try:
            locale.setlocale(locale.LC_ALL, 'Spanish_Spain.1252')
        except:
            pass

     self.manager.cursor.execute('''
        CREATE TABLE IF NOT EXISTS historial_ventas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_hoja TEXT NOT NULL,
            fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            total_venta REAL NOT NULL
        )
    ''')

     self.manager.cursor.execute('''
        CREATE TABLE IF NOT EXISTS detalle_ventas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            venta_id INTEGER,
            codigo_producto TEXT NOT NULL,
            cantidad INTEGER NOT NULL,
            precio_unitario REAL NOT NULL,
            subtotal REAL NOT NULL,
            FOREIGN KEY (venta_id) REFERENCES historial_ventas(id)
        )
    ''')

     self.manager.connection.commit()

     dialog = tk.Toplevel(self.root)
     dialog.title("Realizar Venta")
     dialog.geometry("900x800")  # Increased height for calendar
     dialog.transient(self.root)
     dialog.grab_set()

     main_frame = ttk.Frame(dialog, padding="20")
     main_frame.pack(fill=tk.BOTH, expand=True)

    # Frame para número de hoja y calendario
     sale_frame = ttk.LabelFrame(main_frame, text="Información de Venta", padding="10")
     sale_frame.pack(fill=tk.X, pady=(0, 10))

    # Número de hoja
     ttk.Label(sale_frame, text="Número de Hoja:").pack(side=tk.LEFT, padx=5)
     sale_number = ttk.Entry(sale_frame, width=30)
     sale_number.pack(side=tk.LEFT, padx=5)

    # Calendario
     cal_frame = ttk.Frame(sale_frame)
     cal_frame.pack(side=tk.RIGHT, padx=5)
     ttk.Label(cal_frame, text="Fecha de Venta:").pack()
     cal = Calendar(cal_frame, selectmode='day', date_pattern='yyyy-mm-dd',
                  locale='es_ES',
                  weekenddays=[6,7],
                  weekendbackground='white',
                  weekendforeground='black')
     cal.pack(pady=5)

    # Frame para productos
     products_frame = ttk.LabelFrame(main_frame, text="Productos", padding="10")
     products_frame.pack(fill=tk.BOTH, expand=True)

     columns = ("Código", "Nombre", "Precio", "Cantidad", "Subtotal")
     product_table = ttk.Treeview(products_frame, columns=columns, show="headings")

     for col in columns:
        product_table.heading(col, text=col)
        product_table.column(col, width=100)

     product_table.pack(fill=tk.BOTH, expand=True, pady=5)

    # Frame para entrada de productos
     entry_frame = ttk.Frame(products_frame)
     entry_frame.pack(fill=tk.X, pady=5)

     ttk.Label(entry_frame, text="Código del Producto:").pack(side=tk.LEFT)
     code_entry = ttk.Entry(entry_frame, width=30)
     code_entry.pack(side=tk.LEFT, padx=5)

     self.total_quantity = 0
     self.total_price = 0.0

     total_label = ttk.Label(products_frame, text="Total en cantidad: 0", font=("Arial", 12, "bold"))
     total_label.pack(side=tk.RIGHT, padx=10, pady=5)

     def update_total():
        total_label.config(text=f"Total en cantidad: {self.total_quantity} - Total: Q{self.total_price:.2f}")

     def add_product():
        code = code_entry.get().strip()
        if not code:
            messagebox.showwarning("Advertencia", "Ingrese un código de producto")
            return

        try:
            self.manager.cursor.execute('''SELECT nombre, precio_venta, cantidad FROM productos WHERE codigo = ?''', (code,))
            result = self.manager.cursor.fetchone()
            if result:
                name, price, stock = result
                if stock <= 0:
                    messagebox.showerror("Error", "Stock insuficiente")
                    return
                product_table.insert("", tk.END, values=(code, name, f"Q{price:.2f}", 1, f"Q{price:.2f}"))
                self.total_quantity += 1
                self.total_price += price
                update_total()
                code_entry.delete(0, tk.END)
            else:
                messagebox.showerror("Error", "Producto no encontrado")
        except Exception as e:
            messagebox.showerror("Error", f"Error al buscar el producto: {str(e)}")

     def delete_product():
        selected_item = product_table.selection()
        if not selected_item:
            messagebox.showwarning("Advertencia", "Seleccione un producto para eliminar")
            return
        for item in selected_item:
            values = product_table.item(item, "values")
            if values:
                self.total_quantity -= int(values[3])
                self.total_price -= float(values[4].replace('Q', ''))
                product_table.delete(item)
        update_total()

     def process_sale():
        if not product_table.get_children():
            messagebox.showwarning("Advertencia", "Agregue al menos un producto a la venta")
            return

        hoja_venta = sale_number.get().strip()
        # Obtener la fecha seleccionada del calendario
        sale_date = cal.get_date()
        
        if not hoja_venta:
            messagebox.showwarning("Advertencia", "Ingrese el número de hoja de venta")
            return

        try:
            self.manager.cursor.execute('''INSERT INTO historial_ventas (numero_hoja, fecha, total_venta) VALUES (?, ?, ?)''', 
                                      (hoja_venta, sale_date, self.total_price))
            venta_id = self.manager.cursor.lastrowid

            for item in product_table.get_children():
                values = product_table.item(item, "values")
                code, _, price, quantity, subtotal = values
                quantity = int(quantity)
                price = float(price.replace('Q', ''))
                subtotal = float(subtotal.replace('Q', ''))

                self.manager.cursor.execute('''INSERT INTO detalle_ventas (venta_id, codigo_producto, cantidad, precio_unitario, subtotal) VALUES (?, ?, ?, ?, ?)''', 
                                          (venta_id, code, quantity, price, subtotal))
                self.manager.cursor.execute('''UPDATE productos SET cantidad = cantidad - ? WHERE codigo = ?''', 
                                          (quantity, code))

            self.manager.connection.commit()
            self.load_inventory_data()
            messagebox.showinfo("Éxito", "Venta realizada correctamente")
            dialog.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar la venta: {str(e)}")

     code_entry.bind('<Return>', lambda e: add_product())

    # Botones
     button_frame = ttk.Frame(products_frame)
     button_frame.pack(fill=tk.X, pady=10)

     ttk.Button(button_frame, text="Agregar Producto", command=add_product).pack(side=tk.LEFT, padx=5)
     ttk.Button(button_frame, text="Eliminar Producto", command=delete_product).pack(side=tk.LEFT, padx=5)
     ttk.Button(button_frame, text="Procesar Venta", command=process_sale).pack(side=tk.RIGHT, padx=5)

     code_entry.focus_set()


    
    # [Rest of the previous code remains the same] 
    def setup_ui(self):
    
    # Configuración de la ventana principal
     
     self.root.state('zoomed')
     
     # Agregar después del menú de ventas en setup_ui()
     
     
     
    # Barra de menú
     menubar = tk.Menu(self.root)
     self.root.config(menu=menubar)
    
    # Menú de Base de Datos
     db_menu = tk.Menu(menubar, tearoff=0)
     menubar.add_cascade(label="📂Base de Datos", menu=db_menu)
     db_menu.add_command(label="Nueva Base de Datos", command=self.create_new_db)
     db_menu.add_command(label="Abrir Base de Datos", command=self.open_db)
     db_menu.add_separator()
     db_menu.add_command(label="Realizar Backup", command=self.backup_db)
     db_menu.add_command(label="Restaurar Backup", command=self.restore_db)
     db_menu.add_separator()
     db_menu.add_command(label="Exportar a Excel", command=self.export_to_excel)
     db_menu.add_separator()
     db_menu.add_command(label="Salir", command=self.root.quit)
     
     # Menú de Historial
     

    # AGREGAR AQUÍ EL NUEVO MENÚ DE VENTAS
     sales_menu = tk.Menu(menubar, tearoff=0)
     menubar.add_cascade(label="🛒Ventas", menu=sales_menu)
     sales_menu.add_command(label="Realizar Venta", command=self.make_sale)
     sales_menu.add_command(label="Historial Ventas", command=self.view_sales_history)
     
     # Agregar después del menú de ventas en setup_ui()
     purchase_menu = tk.Menu(menubar, tearoff=0)
     menubar.add_cascade(label="📦Compras", menu=purchase_menu)
     purchase_menu.add_command(label="Realizar Compra", command=self.make_purchase)
     purchase_menu.add_command(label="Ver Historial de Compras", command=self.view_purchase_history)
     # Nuevo Menú de Devoluciones
     returns_menu = tk.Menu(menubar, tearoff=0)
     menubar.add_cascade(label="🔄Devoluciones", menu=returns_menu)
     returns_menu.add_command(label="Realizar Devolución", command=self.make_return)
     returns_menu.add_command(label="Ver Historial de Devoluciones", command=self.view_return_history)

     cuadre_menu = tk.Menu(menubar, tearoff=0)
     menubar.add_cascade(label="📊 Cuadre", menu=cuadre_menu)
     cuadre_menu.add_command(label="Cuadre de Inventario", command=self.cuadre)
     cuadre_menu.add_command(label="Cerrar Trimestre", command=self.cerrar_trimestre)
    
    # Frame principal
     self.main_frame = ttk.Frame(self.root, padding="10")
     self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        # Configurar peso de las filas y columnas
     self.root.grid_rowconfigure(0, weight=1)
     self.root.grid_columnconfigure(0, weight=1)
     self.main_frame.grid_rowconfigure(1, weight=1)
     self.main_frame.grid_columnconfigure(0, weight=1)
        
        # Etiqueta de base de datos actual
     self.db_label = ttk.Label(self.main_frame, text="No hay base de datos seleccionada",
                                 font=('Arial', 12, 'bold'))
     self.db_label.grid(row=0, column=0, pady=10)

        # Frame de búsqueda
     search_frame = ttk.Frame(self.main_frame)
     search_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=5)
        
     ttk.Label(search_frame, text="Buscar:").pack(side=tk.LEFT, padx=5)
     self.search_var = tk.StringVar()
     self.search_var.trace('w', self.search_products)
     ttk.Entry(search_frame, textvariable=self.search_var, width=40).pack(side=tk.LEFT, padx=5)

     ttk.Button(search_frame, text="Actualizar Tabla", command=self.load_inventory_data).pack(side=tk.LEFT, padx=5)

        # Contenedor para el inventario actual
     self.inventory_frame = ttk.Frame(self.main_frame)
     self.inventory_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
     self.inventory_frame.grid_rowconfigure(0, weight=1)
     self.inventory_frame.grid_columnconfigure(0, weight=1)

    def create_new_db(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Nueva Base de Datos")
        dialog.geometry("400x150")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Nombre de la nueva base de datos:",
                 font=('Arial', 10)).pack(pady=10)
        entry = ttk.Entry(frame, width=40)
        entry.pack(pady=5)
        
        def confirm():
            name = entry.get().strip()
            if name:
                if self.manager.create_connection(name):
                    self.update_current_db_display(name)
                    dialog.destroy()
                    self.setup_inventory_view()
                    messagebox.showinfo("Éxito", f"Base de datos '{name}' creada correctamente")
        
        ttk.Button(frame, text="Crear", command=confirm).pack(pady=10)

    def open_db(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Abrir Base de Datos")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Seleccione una base de datos:",
                 font=('Arial', 10)).pack(pady=10)
        
        # Lista de bases de datos
        databases = self.manager.get_available_databases()
        if not databases:
            ttk.Label(frame, text="No hay bases de datos disponibles").pack(pady=5)
            return
        
        listbox = tk.Listbox(frame, width=40, height=10)
        listbox.pack(pady=5)
        
        for db in databases:
            listbox.insert(tk.END, db)
        
        def confirm():
            selection = listbox.curselection()
            if selection:
                db_name = listbox.get(selection[0])
                if self.manager.create_connection(db_name):
                    self.update_current_db_display(db_name)
                    dialog.destroy()
                    self.setup_inventory_view()
        
        ttk.Button(frame, text="Abrir", command=confirm).pack(pady=10)

    def backup_db(self):
        if not self.manager.current_db:
            messagebox.showerror("Error", "No hay una base de datos seleccionada")
            return
        
        if self.manager.backup_database():
            messagebox.showinfo("Éxito", "Backup creado correctamente")

    def restore_db(self):
        if not os.path.exists('backups'):
            messagebox.showerror("Error", "No hay backups disponibles")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Restaurar Backup")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Seleccione un backup para restaurar:",
                 font=('Arial', 10)).pack(pady=10)
        
        listbox = tk.Listbox(frame, width=40, height=10)
        listbox.pack(pady=5)
        
        backups = [f for f in os.listdir('backups') if f.endswith('.db')]
        for backup in backups:
            listbox.insert(tk.END, backup)
        
        def confirm():
            selection = listbox.curselection()
            if selection:
                backup_file = listbox.get(selection[0])
                if messagebox.askyesno("Confirmar", "¿Está seguro de restaurar este backup?"):
                    if self.manager.restore_database(backup_file):
                        self.update_current_db_display(backup_file.split('_backup_')[0])
                        dialog.destroy()
                        self.setup_inventory_view()
                        messagebox.showinfo("Éxito", "Base de datos restaurada correctamente")
        
        ttk.Button(frame, text="Restaurar", command=confirm).pack(pady=10)

    def export_to_excel(self):
        if not self.manager.current_db:
            messagebox.showerror("Error", "No hay una base de datos seleccionada")
            return
        
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"inventario_{self.manager.current_db}"
            )
            
            if not filename:
                return
            
            # Crear un nuevo libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventario"
            
            # Estilos
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Encabezados
            headers = ['Código', 'Nombre', 'Precio Costo', 'Precio Venta', 
                      'Cantidad', 'Categoría', 'Descripción', 'Última Actualización']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
            
            # Datos
            self.manager.cursor.execute('''
                SELECT codigo, nombre, precio_costo, precio_venta, cantidad,
                       categoria, descripcion, fecha_actualizacion
                FROM productos
                ORDER BY nombre
            ''')
            
            for row, product in enumerate(self.manager.cursor.fetchall(), 2):
                for col, value in enumerate(product, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    
                    # Formato para precios
                    if col in [3, 4]:  # Columnas de precios
                        cell.number_format = '"Q"#,##0.00'
            
            # Ajustar anchos de columna
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(filename)
            messagebox.showinfo("Éxito", "Archivo Excel creado correctamente")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel: {str(e)}")

    def update_current_db_display(self, db_name):
        self.db_label.config(text=f"Base de datos actual: {db_name}")

    def setup_inventory_view(self):
        # Limpiar frame de inventario
        for widget in self.inventory_frame.winfo_children():
            widget.destroy()
        
        # Crear vista de inventario
        columns = ('Código', 'Nombre', 'Precio Costo', 'Precio Venta', 
                  'Cantidad', 'Categoría', 'Descripción')
        self.tree = ttk.Treeview(self.inventory_frame, columns=columns, show='headings')
        
        # Configurar columnas
        column_widths = {
            'Código': 100,
            'Nombre': 200,
            'Precio Costo': 100,
            'Precio Venta': 100,
            'Cantidad': 80,
            'Categoría': 100,
            'Descripción': 200
        }
        
        for col in columns:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(c))
            self.tree.column(col, width=column_widths[col])
        
        # Scrollbars
        yscroll = ttk.Scrollbar(self.inventory_frame, orient=tk.VERTICAL, command=self.tree.yview)
        xscroll = ttk.Scrollbar(self.inventory_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        
        # Grid layout
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        yscroll.grid(row=0, column=1, sticky=(tk.N, tk.S))
        xscroll.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Botones de acción
        button_frame = ttk.Frame(self.inventory_frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="Agregar Producto", command=self.add_product).pack(side=tk.LEFT, padx=50)
        ttk.Button(button_frame, text="Editar Producto", command=self.edit_product).pack(side=tk.LEFT, padx=50)
        ttk.Button(button_frame, text="Eliminar Producto", command=self.delete_product).pack(side=tk.LEFT, padx=50)
        
        # Cargar datos
        self.load_inventory_data()

    def sort_treeview(self, col):
     """Ordena el Treeview por la columna seleccionada"""
    
    # Verificar si el Treeview existe
     if not hasattr(self, 'tree') or not self.tree.winfo_exists():
        messagebox.showerror("Error", "No se puede ordenar porque la tabla no existe.")
        return
    
    # Obtener los elementos del Treeview
     items = self.tree.get_children('')
    
     if not items:  # Si no hay elementos, no hacemos nada
        return
    
     try:
        # Intentamos ordenar, convirtiendo los valores numéricos si es necesario
        items = [(self.tree.set(item, col), item) for item in items]
        
        # Determinar si los valores son numéricos o cadenas
        try:
            items.sort(key=lambda x: float(x[0]))  # Intenta ordenar como números
        except ValueError:
            items.sort()  # Si falla, ordena como texto
        
        # Reordenar los elementos en el Treeview
        for index, (val, item) in enumerate(items):
            self.tree.move(item, '', index)
    
     except Exception as e:
        messagebox.showerror("Error", f"Error al ordenar la tabla: {str(e)}")


    def load_inventory_data(self):
        """Carga los datos del inventario en el Treeview"""
        if not self.manager.cursor:
            return
            
        self.tree.delete(*self.tree.get_children())
        self.manager.cursor.execute('''
            SELECT codigo, nombre, precio_costo, precio_venta, cantidad, 
                   categoria, descripcion 
            FROM productos
            ORDER BY nombre
        ''')
        
        for row in self.manager.cursor.fetchall():
            formatted_row = list(row)
            # Formatear precios a 2 decimales
            formatted_row[2] = f"Q{formatted_row[2]:.2f}"
            formatted_row[3] = f"Q{formatted_row[3]:.2f}"
            self.tree.insert('', tk.END, values=formatted_row)

    def search_products(self, *args):
        """Busca productos en tiempo real mientras el usuario escribe"""
        search_term = self.search_var.get().strip().lower()
        
        if not self.manager.cursor:
            return
            
        self.tree.delete(*self.tree.get_children())
        
        if search_term:
            self.manager.cursor.execute('''
                SELECT codigo, nombre, precio_costo, precio_venta, cantidad,
                       categoria, descripcion
                FROM productos
                WHERE LOWER(codigo) LIKE ? OR 
                      LOWER(nombre) LIKE ? OR 
                      LOWER(categoria) LIKE ? OR
                      LOWER(descripcion) LIKE ?
                ORDER BY nombre
            ''', (f'%{search_term}%', f'%{search_term}%', 
                  f'%{search_term}%', f'%{search_term}%'))
        else:
            self.manager.cursor.execute('''
                SELECT codigo, nombre, precio_costo, precio_venta, cantidad,
                       categoria, descripcion
                FROM productos
                ORDER BY nombre
            ''')
        
        for row in self.manager.cursor.fetchall():
            formatted_row = list(row)
            formatted_row[2] = f"Q{formatted_row[2]:.2f}"
            formatted_row[3] = f"Q{formatted_row[3]:.2f}"
            self.tree.insert('', tk.END, values=formatted_row)

    def add_product(self):
     """Muestra la ventana para agregar un nuevo producto"""
     dialog = tk.Toplevel(self.root)
     dialog.title("Agregar Nuevo Producto")
     dialog.geometry("500x600")
     dialog.transient(self.root)
     dialog.grab_set()
    
    # Permitir que Enter guarde los datos
     dialog.bind("<Return>", lambda event: validate_and_save())

     frame = ttk.Frame(dialog, padding="20")
     frame.pack(fill=tk.BOTH, expand=True)
    
    # Campos del formulario
     fields = [
        ("Código*:", "codigo"),
        ("Nombre*:", "nombre"),
        ("Precio Costo*:", "precio_costo"),
        ("Precio Venta*:", "precio_venta"),
        ("Cantidad*:", "cantidad"),
        ("Categoría:", "categoria"),
        ("Descripción:", "descripcion")
    ]
    
     entries = {}
     row = 0
     for label_text, field_name in fields:
        ttk.Label(frame, text=label_text).grid(row=row, column=0, pady=5, sticky=tk.W)
        entry = ttk.Entry(frame, width=40)
        entry.grid(row=row, column=1, pady=5, padx=5)
        entries[field_name] = entry
        row += 1
    
     ttk.Label(frame, text="* Campos obligatorios",
              font=('Arial', 8, 'italic')).grid(row=row, column=0, columnspan=2, pady=10)

     def validate_and_save():
        # Validar campos requeridos
        required_fields = ['codigo', 'nombre', 'precio_costo', 'precio_venta', 'cantidad']
        for field in required_fields:
            if not entries[field].get().strip():
                messagebox.showerror("Error", f"El campo {field} es obligatorio")
                return

        try:
            # Validar valores numéricos
            precio_costo = float(entries['precio_costo'].get())
            precio_venta = float(entries['precio_venta'].get())
            cantidad = int(entries['cantidad'].get())

            if precio_costo < 0 or precio_venta < 0 or cantidad < 0:
                messagebox.showerror("Error", "Los valores numéricos no pueden ser negativos")
                return

            # Insertar en la base de datos
            self.manager.cursor.execute('''
                INSERT INTO productos (codigo, nombre, precio_costo, precio_venta,
                                     cantidad, cantidad_inicial, categoria, descripcion)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                entries['codigo'].get().strip(),
                entries['nombre'].get().strip(),
                precio_costo,
                precio_venta,
                cantidad,
                cantidad,  # Guardar la cantidad inicial
                entries['categoria'].get().strip(),
                entries['descripcion'].get().strip()
            ))

            self.manager.connection.commit()
            dialog.destroy()
            self.load_inventory_data()
            messagebox.showinfo("Éxito", "Producto agregado correctamente")

        except ValueError:
            messagebox.showerror("Error", "Los campos de precio y cantidad deben ser números")
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "Ya existe un producto con ese código")
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar el producto: {str(e)}")

    # BOTÓN VISIBLE PARA GUARDAR
     btn_guardar = ttk.Button(frame, text="Guardar", command=validate_and_save)
     btn_guardar.grid(row=row+1, column=0, columnspan=2, pady=20)

    # Enlazamos la tecla "Enter" al botón también
     dialog.bind("<Return>", lambda event: btn_guardar.invoke())

       

    def edit_product(self):
        """Muestra la ventana para editar un producto seleccionado"""
        if not self.validate_admin_password():
            return
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Por favor, seleccione un producto para editar")
            return
        
        # Obtener datos del producto seleccionado
        item = self.tree.item(selection[0])
        values = item['values']
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Editar Producto")
        dialog.geometry("500x600")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        # Campos del formulario
        fields = [
            ("Código*:", "codigo", values[0]),
            ("Nombre*:", "nombre", values[1]),
            ("Precio Costo*:", "precio_costo", values[2].replace('Q', '')),
            ("Precio Venta*:", "precio_venta", values[3].replace('Q', '')),
            ("Cantidad*:", "cantidad", values[4]),
            ("Categoría:", "categoria", values[5] if values[5] != 'None' else ''),
            ("Descripción:", "descripcion", values[6] if values[6] != 'None' else '')
        ]
        
        entries = {}
        row = 0
        for label_text, field_name, value in fields:
            ttk.Label(frame, text=label_text).grid(row=row, column=0, pady=5, sticky=tk.W)
            entry = ttk.Entry(frame, width=40)
            entry.insert(0, value)
            entry.grid(row=row, column=1, pady=5, padx=5)
            entries[field_name] = entry
            row += 1
        
        ttk.Label(frame, text="* Campos obligatorios",
                 font=('Arial', 8, 'italic')).grid(row=row, column=0, columnspan=2, pady=10)
        
        def validate_and_save():
            # Validar campos requeridos
            required_fields = ['codigo', 'nombre', 'precio_costo', 'precio_venta', 'cantidad']
            for field in required_fields:
                if not entries[field].get().strip():
                    messagebox.showerror("Error", f"El campo {field} es obligatorio")
                    return
            
            try:
                # Validar valores numéricos
                precio_costo = float(entries['precio_costo'].get())
                precio_venta = float(entries['precio_venta'].get())
                cantidad = int(entries['cantidad'].get())
                
                if precio_costo < 0 or precio_venta < 0 or cantidad < 0:
                    messagebox.showerror("Error", "Los valores numéricos no pueden ser negativos")
                    return
                
                # Actualizar en la base de datos
                self.manager.cursor.execute('''
                    UPDATE productos 
                    SET nombre = ?, precio_costo = ?, precio_venta = ?,
                        cantidad = ?, categoria = ?, descripcion = ?
                    WHERE codigo = ?
                ''', (
                    entries['nombre'].get().strip(),
                    precio_costo,
                    precio_venta,
                    cantidad,
                    entries['categoria'].get().strip(),
                    entries['descripcion'].get().strip(),
                    entries['codigo'].get().strip()
                ))
                
                self.manager.connection.commit()
                dialog.destroy()
                self.load_inventory_data()
                messagebox.showinfo("Éxito", "Producto actualizado correctamente")
                
            except ValueError:
                messagebox.showerror("Error", "Los campos de precio y cantidad deben ser números")
            except Exception as e:
                messagebox.showerror("Error", f"Error al actualizar el producto: {str(e)}")
        
        ttk.Button(frame, text="Guardar", command=validate_and_save).grid(
            row=row+1, column=0, columnspan=2, pady=20)
    

    def delete_product(self):
        """Elimina el producto seleccionado"""
        if not self.validate_admin_password():
            return
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Por favor, seleccione un producto para eliminar")
            return
        
        item = self.tree.item(selection[0])
        if messagebox.askyesno("Confirmar", f"¿Está seguro de eliminar el producto '{item['values'][1]}'?"):
            try:
                self.manager.cursor.execute(
                    "DELETE FROM productos WHERE codigo = ?", 
                    (item['values'][0],)
                )
                self.manager.connection.commit()
                self.load_inventory_data()
                messagebox.showinfo("Éxito", "Producto eliminado correctamente")
            except Exception as e:
                messagebox.showerror("Error", f"Error al eliminar el producto: {str(e)}")

def main():
    # Iniciar con la ventana de login en lugar de la aplicación principal
    root = tk.Tk()
    app = LoginApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()